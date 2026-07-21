"""XLSX dry-run and upsert service for CES Calibrations Equipment Lists."""

from __future__ import annotations

import dataclasses
import io
import json
import logging
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from src.domain.exceptions import BadRequestError, ValidationError
from src.domain.models.asset import Asset, AssetCategory, AssetType
from src.domain.models.engineer import Engineer
from src.domain.models.location import Location, LocationKind
from src.domain.models.notification import NotificationChannel, NotificationPriority, NotificationType
from src.domain.models.training_matrix import TrainingMatrixNameMap
from src.domain.models.user import User
from src.domain.services.asset_service import AssetService
from src.domain.services.ces_asset_import_parser import cell_text, normalise_ces_row
from src.domain.services.lookup_similarity import classify_lookup_name, normalise_lookup_key
from src.domain.services.notification_service import NotificationService
from src.domain.services.training_matrix_parser import normalize_person_name

logger = logging.getLogger(__name__)

EQUIPMENT_LIST_SHEET = "Equipment List"


@dataclasses.dataclass(frozen=True)
class RowIssue:
    row: int
    code: str
    message: str
    field: str | None = None
    severity: str = "error"


@dataclasses.dataclass(frozen=True)
class LookupProposal:
    kind: str
    name: str
    intent: str
    reuse_id: int | None
    reuse_name: str | None
    similar_matches: list[dict[str, Any]]
    row_count: int
    needs_confirmation: bool

    def to_dict(self) -> dict[str, Any]:
        return dataclasses.asdict(self)


@dataclasses.dataclass(frozen=True)
class LookupConfirmation:
    kind: str
    name: str
    action: str
    reuse_id: int | None = None


@dataclasses.dataclass
class ValidatedCesRow:
    row: int
    action: str
    asset_type_id: int
    asset_number: str
    name: str
    serial_number: str
    status: str
    existing_id: int | None
    make: str | None = None
    model: str | None = None
    owner_user_id: int | None = None
    location_id: int | None = None
    vehicle_reg: str | None = None
    site: str | None = None
    expiry_date: Any = None
    qr_code_data: str | None = None
    metadata_json: dict[str, Any] = dataclasses.field(default_factory=dict)
    create_type_name: str | None = None
    create_location_name: str | None = None

    def payload(self, *, for_update: bool = False) -> dict[str, Any]:
        data = {
            "asset_type_id": self.asset_type_id,
            "asset_number": self.asset_number,
            "name": self.name,
            "make": self.make,
            "model": self.model,
            "serial_number": self.serial_number,
            "owner_user_id": self.owner_user_id,
            "location_id": self.location_id,
            "vehicle_reg": self.vehicle_reg,
            "site": self.site,
            "expiry_date": self.expiry_date,
            "qr_code_data": self.qr_code_data,
            "status": self.status,
            "metadata_json": self.metadata_json or None,
        }
        if for_update:
            data.pop("asset_number", None)
            if self.owner_user_id is None:
                data.pop("owner_user_id", None)
            if self.vehicle_reg:
                data["location_id"] = None
            elif self.location_id is not None:
                data["vehicle_reg"] = None
            else:
                data.pop("location_id", None)
                data.pop("vehicle_reg", None)
            if self.site is None:
                data.pop("site", None)
        return data


@dataclasses.dataclass
class CesImportReport:
    dry_run: bool
    total_rows: int
    valid_rows: int
    error_rows: int
    creates: int
    updates: int
    errors: list[RowIssue] = dataclasses.field(default_factory=list)
    warnings: list[RowIssue] = dataclasses.field(default_factory=list)
    preview: list[dict[str, Any]] = dataclasses.field(default_factory=list)
    lookup_proposals: list[LookupProposal] = dataclasses.field(default_factory=list)
    requires_confirmation: bool = False

    @property
    def ok(self) -> bool:
        return self.total_rows > 0 and self.error_rows == 0

    def to_dict(self) -> dict[str, Any]:
        def issue(item: RowIssue) -> dict[str, Any]:
            return dataclasses.asdict(item)

        return {
            "dry_run": self.dry_run,
            "total_rows": self.total_rows,
            "valid_rows": self.valid_rows,
            "error_rows": self.error_rows,
            "creates": self.creates,
            "updates": self.updates,
            "ok": self.ok,
            "requires_confirmation": self.requires_confirmation,
            "errors": [issue(item) for item in self.errors],
            "warnings": [issue(item) for item in self.warnings],
            "preview": self.preview,
            "lookup_proposals": [item.to_dict() for item in self.lookup_proposals],
        }


@dataclasses.dataclass
class CesImportCommitResult:
    created_count: int
    updated_count: int
    created_asset_ids: list[int]
    updated_asset_ids: list[int]
    provisional_type_ids: list[int]
    provisional_location_ids: list[int]
    report: CesImportReport

    def to_dict(self) -> dict[str, Any]:
        return {
            "created_count": self.created_count,
            "updated_count": self.updated_count,
            "created_asset_ids": self.created_asset_ids,
            "updated_asset_ids": self.updated_asset_ids,
            "provisional_type_ids": self.provisional_type_ids,
            "provisional_location_ids": self.provisional_location_ids,
            "report": self.report.to_dict(),
        }


def parse_confirmations(raw: Any) -> list[LookupConfirmation]:
    if raw is None or raw == "" or raw == []:
        return []
    if isinstance(raw, str):
        raw = json.loads(raw)
    if not isinstance(raw, list):
        raise BadRequestError("confirmations must be a JSON list")
    out: list[LookupConfirmation] = []
    for item in raw:
        if not isinstance(item, dict):
            raise BadRequestError("Each confirmation must be an object")
        kind = str(item.get("kind") or "").strip()
        name = str(item.get("name") or "").strip()
        action = str(item.get("action") or "").strip()
        reuse_id = item.get("reuse_id")
        if kind not in {"asset_type", "location"} or not name or action not in {"reuse", "create"}:
            raise BadRequestError("Invalid confirmation; need kind, name, action=reuse|create")
        if action == "reuse" and not reuse_id:
            raise BadRequestError(f"Confirmation for '{name}' reuse requires reuse_id")
        out.append(
            LookupConfirmation(
                kind=kind,
                name=name,
                action=action,
                reuse_id=int(reuse_id) if reuse_id is not None else None,
            )
        )
    return out


class CesAssetImportService:
    """Validate then upsert CES assets by serial number."""

    def __init__(self, db: AsyncSession, *, asset_service: AssetService | None = None) -> None:
        self.db = db
        self.assets = asset_service or AssetService(db)

    @staticmethod
    def _duplicate_in_file(
        serial: str,
        equipment_type: str,
        qr: str | None,
        same_file: list[dict[str, Any]],
        row_number: int,
    ) -> RowIssue | None:
        if len(same_file) <= 1:
            return None
        discriminator = (equipment_type.lower(), (qr or "").strip())
        matches = sum(
            (
                str(candidate.get("equipment_type") or "").strip().lower(),
                (candidate.get("qr_code_data") or "").strip(),
            )
            == discriminator
            for candidate in same_file
        )
        if matches <= 1:
            return None
        return RowIssue(
            row_number,
            "AMBIGUOUS_SERIAL",
            f"Serial '{serial}' is repeated in this file without a unique equipment type/QR discriminator",
            "serial_number",
        )

    @staticmethod
    def _match_existing(
        serial: str,
        asset_type: AssetType | None,
        qr: str | None,
        candidates: list[Asset],
        row_number: int,
        claimed_existing: dict[int, int],
    ) -> tuple[Asset | None, list[RowIssue]]:
        issues: list[RowIssue] = []
        # Unresolved / provisional type must not absorb an existing serial of another type.
        if asset_type is None:
            return None, issues
        matched = [asset for asset in candidates if asset.asset_type_id == asset_type.id]
        if qr and matched:
            by_qr = [asset for asset in matched if (asset.qr_code_data or "").strip() == qr.strip()]
            if len(matched) > 1 or by_qr:
                if by_qr:
                    matched = by_qr
                elif len(matched) > 1:
                    issues.append(
                        RowIssue(
                            row_number,
                            "AMBIGUOUS_SERIAL",
                            (
                                f"Serial '{serial}' matches multiple existing assets; "
                                "equipment type and QR must identify one"
                            ),
                            "serial_number",
                        )
                    )
                    return None, issues
        elif len(matched) > 1:
            issues.append(
                RowIssue(
                    row_number,
                    "AMBIGUOUS_SERIAL",
                    (f"Serial '{serial}' matches multiple existing assets; " "equipment type and QR must identify one"),
                    "serial_number",
                )
            )
            return None, issues
        existing = matched[0] if len(matched) == 1 else None
        if existing is not None and existing.id in claimed_existing:
            prior_row = claimed_existing[existing.id]
            message = (
                f"Serial {serial!r} maps to the same register asset ({existing.id}) as "
                f"row {prior_row}; equipment type/QR must identify distinct assets"
            )
            issues.append(RowIssue(row_number, "AMBIGUOUS_SERIAL", message, "serial_number"))
            return None, issues
        return existing, issues

    @staticmethod
    def _confirmation_map(confirmations: list[LookupConfirmation]) -> dict[tuple[str, str], LookupConfirmation]:
        return {(item.kind, normalise_lookup_key(item.name)): item for item in confirmations}

    @staticmethod
    def parse_workbook(content: bytes) -> list[dict[str, Any]]:
        if not content:
            raise BadRequestError("XLSX file is empty")
        if len(content) > 5 * 1024 * 1024:
            raise BadRequestError("XLSX file exceeds 5 MiB limit")
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise BadRequestError(f"Unable to read XLSX workbook: {exc}") from exc
        if EQUIPMENT_LIST_SHEET not in workbook.sheetnames:
            raise BadRequestError(f"Workbook missing required sheet '{EQUIPMENT_LIST_SHEET}'")
        sheet = workbook[EQUIPMENT_LIST_SHEET]
        parsed: list[dict[str, Any]] = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not values or all(value is None or cell_text(value) == "" for value in values):
                continue
            padded = list(values) + [None] * max(0, 12 - len(values))
            raw = {
                "__row__": row_number,
                "location": padded[0],
                "equipment_type": padded[1],
                "make": padded[2],
                "model": padded[3],
                "capacity": padded[4],
                "serial_number": padded[5],
                "asset_id": padded[6],
                "qr_code": padded[7],
                "last_inspection": padded[9],
                "next_inspection": padded[10],
                "status": padded[11],
            }
            try:
                parsed.append(normalise_ces_row(raw))
            except ValueError as exc:
                parsed.append({"__row__": row_number, "__parse_error__": str(exc)})
        if not parsed:
            raise BadRequestError("Equipment List sheet contains no data rows")
        return parsed

    async def _lookups(
        self, tenant_id: int, serials: set[str]
    ) -> tuple[list[AssetType], list[Location], dict[str, int | None], dict[str, list[Asset]]]:
        # Active approved + pending provisional (so re-imports reuse pending rows).
        types = list(
            (
                await self.db.execute(
                    select(AssetType).where(
                        or_(AssetType.tenant_id == tenant_id, AssetType.tenant_id.is_(None)),
                        or_(
                            AssetType.is_active.is_(True),
                            AssetType.approval_status == "pending",
                        ),
                    )
                )
            )
            .scalars()
            .all()
        )
        locations = list(
            (
                await self.db.execute(
                    select(Location).where(
                        Location.tenant_id == tenant_id,
                        or_(
                            Location.is_active.is_(True),
                            Location.approval_status == "pending",
                        ),
                    )
                )
            )
            .scalars()
            .all()
        )
        engineers = (
            await self.db.execute(
                select(Engineer.id, Engineer.display_name, Engineer.user_id).where(Engineer.tenant_id == tenant_id)
            )
        ).all()
        maps = (
            await self.db.execute(
                select(TrainingMatrixNameMap.atlas_name, TrainingMatrixNameMap.engineer_id).where(
                    TrainingMatrixNameMap.tenant_id == tenant_id
                )
            )
        ).all()
        engineer_by_id = {int(engineer_id): user_id for engineer_id, _, user_id in engineers}
        owner_by_name: dict[str, int | None] = {}
        ambiguous: set[str] = set()
        for engineer_id, display_name, user_id in engineers:
            if not display_name:
                continue
            key = normalize_person_name(display_name).lower()
            if key in owner_by_name:
                owner_by_name.pop(key, None)
                ambiguous.add(key)
            elif key not in ambiguous:
                owner_by_name[key] = user_id
        for atlas_name, engineer_id in maps:
            key = normalize_person_name(atlas_name).lower()
            if key not in ambiguous and int(engineer_id) in engineer_by_id:
                owner_by_name[key] = engineer_by_id[int(engineer_id)]

        existing_by_serial: dict[str, list[Asset]] = {}
        if serials:
            assets = (
                (
                    await self.db.execute(
                        select(Asset).where(
                            or_(Asset.tenant_id == tenant_id, Asset.tenant_id.is_(None)),
                            Asset.serial_number.in_(serials),
                        )
                    )
                )
                .scalars()
                .all()
            )
            for asset in assets:
                existing_by_serial.setdefault((asset.serial_number or "").strip(), []).append(asset)
        return types, locations, owner_by_name, existing_by_serial

    def _build_proposals(
        self,
        rows: list[dict[str, Any]],
        types: list[AssetType],
        locations: list[Location],
    ) -> list[LookupProposal]:
        type_candidates = [(item.id, item.name) for item in types]
        location_candidates = [(item.id, item.name) for item in locations]
        type_counts: dict[str, tuple[str, int]] = {}
        location_counts: dict[str, tuple[str, int]] = {}
        for row in rows:
            if "__parse_error__" in row:
                continue
            type_name = str(row.get("equipment_type") or "").strip()
            if type_name:
                key = normalise_lookup_key(type_name)
                display, count = type_counts.get(key, (type_name, 0))
                type_counts[key] = (display, count + 1)
            if row.get("vehicle_reg"):
                continue
            assignment = str(row.get("assignment_text") or "").strip()
            if assignment:
                key = normalise_lookup_key(assignment)
                display, count = location_counts.get(key, (assignment, 0))
                location_counts[key] = (display, count + 1)

        proposals: list[LookupProposal] = []
        for _key, (name, count) in sorted(type_counts.items(), key=lambda item: item[1][0].lower()):
            intent, exact, similar = classify_lookup_name(name, type_candidates)
            proposals.append(
                LookupProposal(
                    kind="asset_type",
                    name=name,
                    intent=intent,
                    reuse_id=exact[0] if exact else None,
                    reuse_name=exact[1] if exact else None,
                    similar_matches=[{"id": m.id, "name": m.name, "score": m.score} for m in similar],
                    row_count=count,
                    needs_confirmation=intent == "similar",
                )
            )
        for _key, (name, count) in sorted(location_counts.items(), key=lambda item: item[1][0].lower()):
            intent, exact, similar = classify_lookup_name(name, location_candidates)
            proposals.append(
                LookupProposal(
                    kind="location",
                    name=name,
                    intent=intent,
                    reuse_id=exact[0] if exact else None,
                    reuse_name=exact[1] if exact else None,
                    similar_matches=[{"id": m.id, "name": m.name, "score": m.score} for m in similar],
                    row_count=count,
                    needs_confirmation=intent == "similar",
                )
            )
        return proposals

    def _resolve_lookup(
        self,
        *,
        kind: str,
        name: str,
        proposals_by_key: dict[tuple[str, str], LookupProposal],
        confirmations: dict[tuple[str, str], LookupConfirmation],
        by_id: dict[int, Any],
        enforce_confirmations: bool,
    ) -> tuple[int | None, str | None, list[RowIssue]]:
        """Return (existing_id, create_name, issues)."""
        key = (kind, normalise_lookup_key(name))
        proposal = proposals_by_key.get(key)
        if proposal is None:
            return None, None, []
        if proposal.intent == "reuse" and proposal.reuse_id is not None:
            return proposal.reuse_id, None, []
        confirmation = confirmations.get(key)
        if proposal.intent == "similar":
            if not confirmation:
                if enforce_confirmations:
                    return (
                        None,
                        None,
                        [
                            RowIssue(
                                0,
                                "NEEDS_CONFIRMATION",
                                (
                                    f"{kind} '{name}' is similar to existing "
                                    f"{[m['name'] for m in proposal.similar_matches]}; confirm reuse or create"
                                ),
                                kind,
                            )
                        ],
                    )
                # Dry-run: treat as pending create for validity preview.
                return None, name, []
            if confirmation.action == "reuse":
                if confirmation.reuse_id not in by_id:
                    return (
                        None,
                        None,
                        [RowIssue(0, "INVALID_CONFIRMATION", f"reuse_id {confirmation.reuse_id} not found", kind)],
                    )
                return confirmation.reuse_id, None, []
            return None, name, []
        # new
        return None, name, []

    @staticmethod
    def _missing_confirmation_errors(
        proposals: list[LookupProposal],
        confirmation_map: dict[tuple[str, str], LookupConfirmation],
    ) -> list[RowIssue]:
        errors: list[RowIssue] = []
        for proposal in proposals:
            if not proposal.needs_confirmation:
                continue
            key = (proposal.kind, normalise_lookup_key(proposal.name))
            if key in confirmation_map:
                continue
            errors.append(
                RowIssue(
                    0,
                    "NEEDS_CONFIRMATION",
                    (
                        f"{proposal.kind} '{proposal.name}' looks similar to "
                        f"{[m['name'] for m in proposal.similar_matches]}; confirm before commit"
                    ),
                    proposal.kind,
                )
            )
        return errors

    @staticmethod
    def _lookup_intent_warnings(
        *,
        proposal: LookupProposal | None,
        row_number: int,
        display_name: str,
        field: str,
        similar_code: str,
        new_code: str,
        enforce_confirmations: bool,
    ) -> list[RowIssue]:
        if proposal is None:
            return []
        if proposal.intent == "similar" and not enforce_confirmations:
            return [
                RowIssue(
                    row_number,
                    similar_code,
                    (
                        f"{field.replace('_', ' ').title()} '{display_name}' is similar to "
                        f"{[m['name'] for m in proposal.similar_matches]}; confirm before commit"
                    ),
                    field,
                    "warning",
                )
            ]
        if proposal.intent == "new":
            return [
                RowIssue(
                    row_number,
                    new_code,
                    f"{field.replace('_', ' ').title()} '{display_name}' will be created pending approval",
                    field,
                    "warning",
                )
            ]
        return []

    def _validate_one_row(
        self,
        row: dict[str, Any],
        *,
        proposals_by_key: dict[tuple[str, str], LookupProposal],
        confirmation_map: dict[tuple[str, str], LookupConfirmation],
        type_by_id: dict[int, AssetType],
        location_by_id: dict[int, Location],
        owner_by_name: dict[str, int | None],
        existing_by_serial: dict[str, list[Asset]],
        serial_rows: dict[str, list[dict[str, Any]]],
        claimed_existing: dict[int, int],
        enforce_confirmations: bool,
    ) -> tuple[list[RowIssue], list[RowIssue], ValidatedCesRow | None]:
        row_number = int(row["__row__"])
        row_errors: list[RowIssue] = []
        warnings: list[RowIssue] = []
        if "__parse_error__" in row:
            return [RowIssue(row_number, "INVALID_ROW", row["__parse_error__"])], [], None

        serial = str(row["serial_number"]).strip()
        equipment_type = str(row["equipment_type"]).strip()
        qr = row["qr_code_data"]
        if not serial:
            row_errors.append(RowIssue(row_number, "REQUIRED", "Serial Number is required", "serial_number"))
        if not equipment_type:
            row_errors.append(RowIssue(row_number, "REQUIRED", "Equipment Type is required", "equipment_type"))
        if not row["name"]:
            row_errors.append(RowIssue(row_number, "REQUIRED", "Equipment name is required", "equipment_type"))

        type_id: int | None = None
        create_type_name: str | None = None
        if equipment_type:
            type_id, create_type_name, resolve_issues = self._resolve_lookup(
                kind="asset_type",
                name=equipment_type,
                proposals_by_key=proposals_by_key,
                confirmations=confirmation_map,
                by_id=type_by_id,
                enforce_confirmations=enforce_confirmations,
            )
            for issue in resolve_issues:
                row_errors.append(RowIssue(row_number, issue.code, issue.message, issue.field, issue.severity))
            warnings.extend(
                self._lookup_intent_warnings(
                    proposal=proposals_by_key.get(("asset_type", normalise_lookup_key(equipment_type))),
                    row_number=row_number,
                    display_name=equipment_type,
                    field="equipment_type",
                    similar_code="SIMILAR_TYPE",
                    new_code="TYPE_WILL_CREATE_PENDING",
                    enforce_confirmations=enforce_confirmations,
                )
            )

        asset_type = type_by_id.get(type_id) if type_id else None
        duplicate = self._duplicate_in_file(serial, equipment_type, qr, serial_rows.get(serial, []), row_number)
        if duplicate:
            row_errors.append(duplicate)

        existing, match_issues = self._match_existing(
            serial,
            asset_type,
            qr,
            list(existing_by_serial.get(serial, [])),
            row_number,
            claimed_existing,
        )
        row_errors.extend(match_issues)

        location_id: int | None = None
        create_location_name: str | None = None
        owner_user_id = None
        vehicle_reg = row["vehicle_reg"]
        assignment = str(row.get("assignment_text") or "").strip()
        if not vehicle_reg and assignment:
            location_id, create_location_name, loc_issues = self._resolve_lookup(
                kind="location",
                name=assignment,
                proposals_by_key=proposals_by_key,
                confirmations=confirmation_map,
                by_id=location_by_id,
                enforce_confirmations=enforce_confirmations,
            )
            for issue in loc_issues:
                row_errors.append(RowIssue(row_number, issue.code, issue.message, issue.field, issue.severity))
            warnings.extend(
                self._lookup_intent_warnings(
                    proposal=proposals_by_key.get(("location", normalise_lookup_key(assignment))),
                    row_number=row_number,
                    display_name=assignment,
                    field="location",
                    similar_code="SIMILAR_LOCATION",
                    new_code="LOCATION_WILL_CREATE_PENDING",
                    enforce_confirmations=enforce_confirmations,
                )
            )

        owner_name = row["engineer_name"]
        if owner_name:
            key = normalize_person_name(owner_name).lower()
            if key in owner_by_name and owner_by_name[key] is not None:
                owner_user_id = owner_by_name[key]
            else:
                warnings.append(
                    RowIssue(
                        row_number,
                        "UNMAPPED_OWNER",
                        f"Engineer '{owner_name}' has no uniquely linked user",
                        "location",
                        "warning",
                    )
                )

        if row["not_made_available"]:
            warnings.append(
                RowIssue(
                    row_number,
                    "NOT_MADE_AVAILABLE",
                    "CES status is Not Made Available; imported as active and flagged in metadata",
                    "status",
                    "warning",
                )
            )
        if row_errors:
            return row_errors, warnings, None
        if type_id is None and not create_type_name:
            return (
                [RowIssue(row_number, "UNKNOWN_TYPE", f"Unknown asset type '{equipment_type}'", "equipment_type")],
                warnings,
                None,
            )

        action = "update" if existing else "create"
        if existing is not None:
            claimed_existing[existing.id] = row_number
        metadata = {
            "ces_location": row["location_raw"],
            "ces_company": row["company"],
            "ces_last_inspection": row["last_inspection"].isoformat() if row["last_inspection"] else None,
            "ces_asset_id": row["asset_id"],
            "ces_not_made_available": row["not_made_available"],
        }
        validated = ValidatedCesRow(
            row=row_number,
            action=action,
            asset_type_id=type_id or 0,
            asset_number=serial or qr or f"CES-{row_number}",
            name=row["name"][:200],
            serial_number=serial,
            status=row["status"],
            existing_id=existing.id if existing else None,
            make=row["make"],
            model=row["model"],
            owner_user_id=owner_user_id,
            location_id=location_id,
            vehicle_reg=vehicle_reg,
            site=row["assignment_text"],
            expiry_date=row["expiry_date"],
            qr_code_data=qr,
            metadata_json={key: value for key, value in metadata.items() if value is not None},
            create_type_name=create_type_name,
            create_location_name=create_location_name,
        )
        return [], warnings, validated

    async def validate_rows(
        self,
        rows: list[dict[str, Any]],
        *,
        tenant_id: int,
        dry_run: bool,
        confirmations: list[LookupConfirmation] | None = None,
        enforce_confirmations: bool = False,
    ) -> tuple[CesImportReport, list[ValidatedCesRow]]:
        serials = {str(row.get("serial_number") or "").strip() for row in rows if row.get("serial_number")}
        types, locations, owner_by_name, existing_by_serial = await self._lookups(tenant_id, serials)
        type_by_id = {item.id: item for item in types}
        location_by_id = {item.id: item for item in locations}
        proposals = self._build_proposals(rows, types, locations)
        proposals_by_key = {(p.kind, normalise_lookup_key(p.name)): p for p in proposals}
        confirmation_map = self._confirmation_map(confirmations or [])

        errors: list[RowIssue] = []
        if enforce_confirmations:
            errors.extend(self._missing_confirmation_errors(proposals, confirmation_map))

        serial_rows: dict[str, list[dict[str, Any]]] = {}
        for row in rows:
            serial = str(row.get("serial_number") or "").strip()
            if serial:
                serial_rows.setdefault(serial, []).append(row)

        warnings: list[RowIssue] = []
        validated: list[ValidatedCesRow] = []
        claimed_existing: dict[int, int] = {}
        creates = updates = 0

        for row in rows:
            row_errors, row_warnings, item = self._validate_one_row(
                row,
                proposals_by_key=proposals_by_key,
                confirmation_map=confirmation_map,
                type_by_id=type_by_id,
                location_by_id=location_by_id,
                owner_by_name=owner_by_name,
                existing_by_serial=existing_by_serial,
                serial_rows=serial_rows,
                claimed_existing=claimed_existing,
                enforce_confirmations=enforce_confirmations,
            )
            errors.extend(row_errors)
            warnings.extend(row_warnings)
            if item is None:
                continue
            creates += item.action == "create"
            updates += item.action == "update"
            validated.append(item)

        preview = [
            {
                "row": item.row,
                "action": item.action,
                "asset_number": item.asset_number,
                "name": item.name,
                "serial_number": item.serial_number,
                "owner_user_id": item.owner_user_id,
                "location_id": item.location_id,
                "vehicle_reg": item.vehicle_reg,
                "status": item.status,
                "not_made_available": item.metadata_json.get("ces_not_made_available", False),
            }
            for item in validated[:50]
        ]
        requires_confirmation = any(p.needs_confirmation for p in proposals)
        report = CesImportReport(
            dry_run=dry_run,
            total_rows=len(rows),
            valid_rows=len(validated),
            error_rows=len({issue.row for issue in errors if issue.row > 0})
            + (1 if any(issue.row == 0 for issue in errors) else 0),
            creates=creates,
            updates=updates,
            errors=errors,
            warnings=warnings,
            preview=preview,
            lookup_proposals=proposals,
            requires_confirmation=requires_confirmation,
        )
        return report, validated

    async def dry_run(self, content: bytes, *, tenant_id: int) -> CesImportReport:
        report, _ = await self.validate_rows(
            self.parse_workbook(content),
            tenant_id=tenant_id,
            dry_run=True,
            enforce_confirmations=False,
        )
        return report

    async def _ensure_provisional_type(self, name: str, *, user_id: int, tenant_id: int) -> AssetType:
        existing = (
            (
                await self.db.execute(
                    select(AssetType).where(
                        or_(AssetType.tenant_id == tenant_id, AssetType.tenant_id.is_(None)),
                        AssetType.approval_status == "pending",
                    )
                )
            )
            .scalars()
            .all()
        )
        for item in existing:
            if normalise_lookup_key(item.name) == normalise_lookup_key(name):
                return item
        asset_type = AssetType(
            category=AssetCategory.SAFETY,
            name=name[:200],
            description="Provisional type from CES import — pending approval",
            is_active=False,
            approval_status="pending",
            source="ces_import",
            tenant_id=tenant_id,
            created_by_id=user_id,
            updated_by_id=user_id,
        )
        self.db.add(asset_type)
        await self.db.flush()
        return asset_type

    async def _ensure_provisional_location(self, name: str, *, user_id: int, tenant_id: int) -> Location:
        existing = (
            (
                await self.db.execute(
                    select(Location).where(
                        Location.tenant_id == tenant_id,
                        Location.approval_status == "pending",
                    )
                )
            )
            .scalars()
            .all()
        )
        for item in existing:
            if normalise_lookup_key(item.name) == normalise_lookup_key(name):
                return item
        location = Location(
            name=name[:200],
            kind=LocationKind.SITE,
            is_active=False,
            approval_status="pending",
            source="ces_import",
            tenant_id=tenant_id,
            created_by_id=user_id,
            updated_by_id=user_id,
        )
        self.db.add(location)
        await self.db.flush()
        return location

    async def _notify_pending_lookups(
        self,
        *,
        user_id: int,
        type_count: int,
        location_count: int,
    ) -> None:
        if type_count + location_count <= 0:
            return
        admins = list(
            (
                await self.db.execute(
                    select(User).where(
                        or_(User.is_superuser.is_(True), User.is_active.is_(True)),
                    )
                )
            )
            .scalars()
            .all()
        )
        # Prefer superusers; otherwise notify active users with admin-like emails is too broad.
        targets = [u for u in admins if getattr(u, "is_superuser", False)]
        if not targets:
            return
        notifier = NotificationService(self.db)
        for admin in targets[:10]:
            if admin.id == user_id:
                continue
            await notifier.create_notification(
                user_id=admin.id,
                notification_type=NotificationType.APPROVAL_REQUESTED,
                title="Safety lookups awaiting approval",
                message=(
                    f"CES import proposed {type_count} asset type(s) and {location_count} location(s) "
                    "for approval in Admin → Lookup Tables."
                ),
                priority=NotificationPriority.HIGH,
                entity_type="safety_lookup_proposal",
                entity_id="pending",
                action_url="/admin/lookups?pending=safety",
                sender_id=user_id,
                channels=[NotificationChannel.IN_APP, NotificationChannel.EMAIL],
            )

    async def commit(
        self,
        content: bytes,
        *,
        user_id: int,
        tenant_id: int,
        confirmations: list[LookupConfirmation] | None = None,
    ) -> CesImportCommitResult:
        report, validated = await self.validate_rows(
            self.parse_workbook(content),
            tenant_id=tenant_id,
            dry_run=False,
            confirmations=confirmations,
            enforce_confirmations=True,
        )
        if not report.ok:
            raise ValidationError(
                "CES import validation failed; fix row errors or confirm similar lookups before commit",
                code="CES_ASSET_IMPORT_VALIDATION_FAILED",
                details=report.to_dict(),
            )
        created_ids: list[int] = []
        updated_ids: list[int] = []
        provisional_type_ids: list[int] = []
        provisional_location_ids: list[int] = []
        type_cache: dict[str, int] = {}
        location_cache: dict[str, int] = {}
        try:
            for item in validated:
                if item.create_type_name:
                    key = normalise_lookup_key(item.create_type_name)
                    if key not in type_cache:
                        created = await self._ensure_provisional_type(
                            item.create_type_name, user_id=user_id, tenant_id=tenant_id
                        )
                        type_cache[key] = created.id
                        provisional_type_ids.append(created.id)
                    item.asset_type_id = type_cache[key]
                if item.create_location_name:
                    key = normalise_lookup_key(item.create_location_name)
                    if key not in location_cache:
                        created = await self._ensure_provisional_location(
                            item.create_location_name, user_id=user_id, tenant_id=tenant_id
                        )
                        location_cache[key] = created.id
                        provisional_location_ids.append(created.id)
                    item.location_id = location_cache[key]

                if item.action == "update":
                    assert item.existing_id is not None
                    existing = await self.assets.get_asset(item.existing_id, tenant_id=tenant_id)
                    update_payload = item.payload(for_update=True)
                    prior_meta = dict(existing.metadata_json or {})
                    ces_meta = dict(item.metadata_json or {})
                    update_payload["metadata_json"] = {**prior_meta, **ces_meta} or None
                    asset = await self.assets.update_asset(
                        item.existing_id,
                        update_payload,
                        tenant_id=tenant_id,
                        actor_user_id=user_id,
                        commit=False,
                    )
                    updated_ids.append(asset.id)
                else:
                    asset = await self.assets.create_asset(
                        item.payload(),
                        user_id=user_id,
                        tenant_id=tenant_id,
                        commit=False,
                    )
                    created_ids.append(asset.id)
            await self.db.commit()
        except Exception:
            await self.db.rollback()
            raise
        # Notify after durable commit so notification failures cannot roll back assets.
        try:
            await self._notify_pending_lookups(
                user_id=user_id,
                type_count=len(set(provisional_type_ids)),
                location_count=len(set(provisional_location_ids)),
            )
        except Exception:  # noqa: BLE001
            logger.exception("CES import succeeded but pending-lookup notification failed")
        report.dry_run = False
        return CesImportCommitResult(
            created_count=len(created_ids),
            updated_count=len(updated_ids),
            created_asset_ids=created_ids,
            updated_asset_ids=updated_ids,
            provisional_type_ids=sorted(set(provisional_type_ids)),
            provisional_location_ids=sorted(set(provisional_location_ids)),
            report=report,
        )

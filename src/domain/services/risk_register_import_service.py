"""Excel bulk import for Enterprise Risk Register (RR-W4 + Action Plan→CAPA).

Dry-run validates the Risk Register sheet and Action Plan sheet (when present).
Commit upserts EnterpriseRisk rows (PELR*) then creates/updates CAPA actions from
Action Plan rows linked by risk reference.
"""

from __future__ import annotations

import dataclasses
import io
import re
from datetime import datetime, timedelta, timezone
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.domain.exceptions import BadRequestError, ValidationError
from src.domain.models.capa import CAPAAction, CAPASource
from src.domain.models.risk_register import EnterpriseRisk, RiskAppetiteStatement
from src.domain.services.risk_service import RiskService, RiskScoringEngine

RISK_REGISTER_SHEET = "Risk Register"
ACTION_PLAN_SHEET = "Action Plan"
REFERENCE_PATTERN = re.compile(r"^PELR\d+$", re.IGNORECASE)

HEADER_ALIASES: dict[str, str] = {
    "ref": "reference",
    "date identified": "identified_date",
    "risk title": "title",
    "risk description": "description",
    "root causes": "root_causes",
    "category": "category",
    "risk owner": "risk_owner_name",
    "gross impact (1-5)": "inherent_impact",
    "gross likelihood (1-5)": "inherent_likelihood",
    "gross score": "inherent_score_raw",
    "gross rag": "inherent_rag",
    "existing controls": "treatment_plan",
    "control effectiveness": "control_effectiveness",
    "net impact (1-5)": "residual_impact",
    "net likelihood (1-5)": "residual_likelihood",
    "net score": "residual_score_raw",
    "net rag": "residual_rag",
    "trend": "trend",
    "status": "status",
    "last reviewed": "last_review_date",
    "next review": "next_review_date",
    "comments": "comments",
    "sortkey": "sort_key",
}

ACTION_PLAN_HEADER_ALIASES: dict[str, str] = {
    "action id": "action_id",
    "linked risk ref": "risk_reference",
    "action description": "description",
    "owner": "owner",
    "cost (gbp)": "cost",
    "deadline": "deadline",
    "status": "status",
    "progress notes": "progress_notes",
    "matchkey": "match_key",
}

CATEGORY_MAP: dict[str, str] = {
    "compliance": "compliance",
    "operational": "operational",
    "financial": "financial",
    "strategic": "strategic",
    "reputational": "reputational",
    "health & safety": "health_safety",
    "health and safety": "health_safety",
    "health_safety": "health_safety",
    "environmental": "environmental",
    "technological": "technological",
    "legal": "legal",
    "project": "project",
}

STATUS_MAP: dict[str, str] = {
    "open": "active",
    "active": "active",
    "closed": "closed",
    "monitoring": "monitoring",
    "mitigated": "mitigated",
    "draft": "draft",
    "identified": "identified",
    "assessing": "assessing",
    "treating": "treating",
    "escalated": "escalated",
}


@dataclasses.dataclass(frozen=True)
class RowError:
    row: int
    code: str
    message: str
    field: str | None = None


@dataclasses.dataclass
class ValidatedImportRow:
    row: int
    action: str  # create | update
    reference: str
    title: str
    description: str
    category: str
    inherent_likelihood: int
    inherent_impact: int
    residual_likelihood: int
    residual_impact: int
    risk_owner_name: str | None = None
    treatment_plan: str | None = None
    status: str = "active"
    identified_date: datetime | None = None
    last_review_date: datetime | None = None
    next_review_date: datetime | None = None
    review_notes: str | None = None
    context: str | None = None
    existing_id: int | None = None


@dataclasses.dataclass
class ValidatedActionPlanRow:
    row: int
    action: str  # create | update
    action_id: str
    risk_reference: str
    title: str
    description: str
    status: str
    match_key: str
    owner: str | None = None
    due_date: datetime | None = None
    progress_notes: str | None = None
    existing_capa_id: int | None = None


@dataclasses.dataclass
class ImportValidationReport:
    dry_run: bool
    total_rows: int
    valid_rows: int
    error_rows: int
    creates: int
    updates: int
    errors: list[RowError]
    preview: list[dict[str, Any]] = dataclasses.field(default_factory=list)
    action_plan_skipped: bool = False
    action_plan_total_rows: int = 0
    action_plan_creates: int = 0
    action_plan_updates: int = 0
    action_plan_error_rows: int = 0
    action_plan_errors: list[RowError] = dataclasses.field(default_factory=list)
    action_plan_preview: list[dict[str, Any]] = dataclasses.field(default_factory=list)

    @property
    def ok(self) -> bool:
        return self.error_rows == 0 and self.total_rows > 0 and self.action_plan_error_rows == 0

    def to_dict(self) -> dict[str, Any]:
        return {
            "dry_run": self.dry_run,
            "total_rows": self.total_rows,
            "valid_rows": self.valid_rows,
            "error_rows": self.error_rows,
            "creates": self.creates,
            "updates": self.updates,
            "ok": self.ok,
            "errors": [
                {
                    "row": e.row,
                    "code": e.code,
                    "message": e.message,
                    "field": e.field,
                }
                for e in self.errors
            ],
            "preview": self.preview,
            "action_plan_skipped": self.action_plan_skipped,
            "action_plan_total_rows": self.action_plan_total_rows,
            "action_plan_creates": self.action_plan_creates,
            "action_plan_updates": self.action_plan_updates,
            "action_plan_error_rows": self.action_plan_error_rows,
            "action_plan_errors": [
                {
                    "row": e.row,
                    "code": e.code,
                    "message": e.message,
                    "field": e.field,
                }
                for e in self.action_plan_errors
            ],
            "action_plan_preview": self.action_plan_preview,
        }


@dataclasses.dataclass
class ImportCommitResult:
    created_count: int
    updated_count: int
    created_risk_ids: list[int]
    updated_risk_ids: list[int]
    report: ImportValidationReport
    capa_created_count: int = 0
    capa_updated_count: int = 0
    capa_created_ids: list[int] = dataclasses.field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "created_count": self.created_count,
            "updated_count": self.updated_count,
            "created_risk_ids": self.created_risk_ids,
            "updated_risk_ids": self.updated_risk_ids,
            "capa_created_count": self.capa_created_count,
            "capa_updated_count": self.capa_updated_count,
            "capa_created_ids": self.capa_created_ids,
            "report": self.report.to_dict(),
        }


class RiskRegisterImportService:
    """Parse, validate, and optionally commit Plantexpand Risk Register XLSX imports."""

    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    @staticmethod
    def _normalise_header(value: Any) -> str:
        return str(value or "").strip().lower()

    @classmethod
    def _parse_sheet_rows(
        cls,
        workbook: Any,
        sheet_name: str,
        aliases: dict[str, str],
        *,
        required_fields: set[str],
    ) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            raise BadRequestError(f"Workbook missing required sheet '{sheet_name}'")

        sheet = workbook[sheet_name]
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise BadRequestError(f"Sheet '{sheet_name}' is empty") from exc

        column_map: dict[int, str] = {}
        for idx, raw_header in enumerate(header_row):
            canonical = aliases.get(cls._normalise_header(raw_header))
            if canonical:
                column_map[idx] = canonical

        missing = required_fields - set(column_map.values())
        if missing:
            raise BadRequestError(f"{sheet_name} sheet missing required column(s): " + ", ".join(sorted(missing)))

        parsed: list[dict[str, Any]] = []
        for row_number, raw_row in enumerate(rows_iter, start=2):
            if raw_row is None or all(v is None or str(v).strip() == "" for v in raw_row):
                continue
            row_data: dict[str, Any] = {"__row__": row_number}
            for idx, field in column_map.items():
                if idx < len(raw_row):
                    row_data[field] = raw_row[idx]
            parsed.append(row_data)
        return parsed

    @classmethod
    def parse_workbook(cls, content: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]] | None]:
        """Return (register_rows, action_plan_rows_or_None).

        ``None`` for Action Plan means the sheet is absent (honest skip).
        Empty Action Plan sheet returns ``[]``.
        """
        if not content:
            raise BadRequestError("XLSX file is empty")
        if len(content) > 5 * 1024 * 1024:
            raise BadRequestError("XLSX file exceeds 5 MiB limit")

        try:
            workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001 — surface as client error
            raise BadRequestError(f"Unable to read XLSX workbook: {exc}") from exc

        register_rows = cls._parse_sheet_rows(
            workbook,
            RISK_REGISTER_SHEET,
            HEADER_ALIASES,
            required_fields={"reference", "title", "description"},
        )
        if not register_rows:
            raise BadRequestError("Risk Register sheet contains no data rows")

        if ACTION_PLAN_SHEET not in workbook.sheetnames:
            return register_rows, None

        action_rows = cls._parse_sheet_rows(
            workbook,
            ACTION_PLAN_SHEET,
            ACTION_PLAN_HEADER_ALIASES,
            required_fields={"action_id", "risk_reference", "description"},
        )
        return register_rows, action_rows

    @classmethod
    def parse_xlsx(cls, content: bytes) -> list[dict[str, Any]]:
        """Backward-compatible: parse Risk Register sheet only."""
        register_rows, _ = cls.parse_workbook(content)
        return register_rows

    @staticmethod
    def _cell_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat()
        return str(value).strip()

    @staticmethod
    def _parse_score(value: Any, *, field: str, row: int, errors: list[RowError]) -> int | None:
        if value is None or str(value).strip() == "":
            errors.append(RowError(row=row, field=field, code="REQUIRED", message=f"{field} is required (1-5)"))
            return None
        try:
            score = int(float(value))
        except (TypeError, ValueError):
            errors.append(
                RowError(
                    row=row,
                    field=field,
                    code="INVALID_SCORE",
                    message=f"{field} must be an integer between 1 and 5",
                )
            )
            return None
        if score < 1 or score > 5:
            errors.append(
                RowError(
                    row=row,
                    field=field,
                    code="INVALID_SCORE",
                    message=f"{field} must be between 1 and 5",
                )
            )
            return None
        return score

    @staticmethod
    def _parse_datetime(value: Any) -> datetime | None:
        if value is None or str(value).strip() == "":
            return None
        text = str(value).strip()
        if text.lower() in {"no deadline", "n/a", "na", "-"}:
            return None
        if isinstance(value, datetime):
            dt = value
        else:
            try:
                dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
            except ValueError:
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                    try:
                        dt = datetime.strptime(text, fmt)
                        break
                    except ValueError:
                        continue
                else:
                    return None
        if dt.tzinfo is not None:
            dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt

    @staticmethod
    def _normalise_category(raw: str) -> str | None:
        key = raw.strip().lower()
        if not key:
            return None
        return CATEGORY_MAP.get(key, key.replace(" ", "_").replace("&", "").replace("__", "_"))

    @staticmethod
    def _normalise_status(raw: str) -> str:
        key = raw.strip().lower()
        return STATUS_MAP.get(key, "active")

    @staticmethod
    def _build_review_notes(comments: str, control_effectiveness: str) -> str | None:
        parts: list[str] = []
        if comments.strip():
            parts.append(comments.strip())
        if control_effectiveness.strip():
            parts.append(f"Control effectiveness: {control_effectiveness.strip()}")
        if not parts:
            return None
        return "\n\n".join(parts)

    @staticmethod
    def _action_title(description: str, action_id: str) -> str:
        first_line = description.strip().splitlines()[0].strip() if description.strip() else ""
        # Strip leading "1. " numbering commonly used in Action Plan rows.
        cleaned = re.sub(r"^\d+\.\s*", "", first_line).strip()
        if cleaned:
            return cleaned[:255]
        return f"Action {action_id}"

    async def _existing_references(self, references: set[str], tenant_id: int) -> dict[str, EnterpriseRisk]:
        if not references:
            return {}
        result = await self.db.execute(
            select(EnterpriseRisk).where(
                EnterpriseRisk.tenant_id == tenant_id,
                EnterpriseRisk.reference.in_(sorted(references)),
            )
        )
        return {risk.reference.upper(): risk for risk in result.scalars().all()}

    async def _existing_capa_by_source_refs(self, source_refs: set[str], tenant_id: int) -> dict[str, CAPAAction]:
        if not source_refs:
            return {}
        result = await self.db.execute(
            select(CAPAAction).where(
                CAPAAction.tenant_id == tenant_id,
                CAPAAction.source_type == CAPASource.RISK,
                CAPAAction.source_reference.in_(sorted(source_refs)),
            )
        )
        return {str(a.source_reference): a for a in result.scalars().all() if a.source_reference}

    async def _appetite_threshold(self, category: str) -> int:
        result = await self.db.execute(select(RiskAppetiteStatement).where(RiskAppetiteStatement.category == category))
        appetite = result.scalar_one_or_none()
        return appetite.max_residual_score if appetite else 12

    async def validate_rows(
        self,
        rows: list[dict[str, Any]],
        *,
        tenant_id: int,
        dry_run: bool = True,
    ) -> tuple[ImportValidationReport, list[ValidatedImportRow]]:
        references = {
            self._cell_text(row.get("reference")).upper() for row in rows if self._cell_text(row.get("reference"))
        }
        existing = await self._existing_references(references, tenant_id)

        errors: list[RowError] = []
        validated: list[ValidatedImportRow] = []
        seen_in_file: dict[str, int] = {}
        creates = 0
        updates = 0

        for row in rows:
            row_number = int(row["__row__"])
            row_errors: list[RowError] = []

            reference = self._cell_text(row.get("reference")).upper()
            title = self._cell_text(row.get("title"))
            description = self._cell_text(row.get("description"))
            root_causes = self._cell_text(row.get("root_causes"))
            category_raw = self._cell_text(row.get("category"))
            owner = self._cell_text(row.get("risk_owner_name")) or None
            treatment_plan = self._cell_text(row.get("treatment_plan")) or None
            comments = self._cell_text(row.get("comments"))
            control_effectiveness = self._cell_text(row.get("control_effectiveness"))
            status_raw = self._cell_text(row.get("status"))

            if not reference:
                row_errors.append(
                    RowError(row=row_number, field="reference", code="REQUIRED", message="Ref is required")
                )
            elif not REFERENCE_PATTERN.match(reference):
                row_errors.append(
                    RowError(
                        row=row_number,
                        field="reference",
                        code="INVALID_REFERENCE",
                        message=f"Reference '{reference}' must match PELR* pattern (e.g. PELR1)",
                    )
                )
            elif reference in seen_in_file:
                row_errors.append(
                    RowError(
                        row=row_number,
                        field="reference",
                        code="DUPLICATE_IN_FILE",
                        message=(f"Duplicate reference '{reference}' " f"(also on row {seen_in_file[reference]})"),
                    )
                )
            else:
                seen_in_file[reference] = row_number

            if not title or len(title) < 5:
                row_errors.append(
                    RowError(
                        row=row_number,
                        field="title",
                        code="INVALID_TITLE",
                        message="Risk Title must be at least 5 characters",
                    )
                )
            if not description or len(description) < 10:
                row_errors.append(
                    RowError(
                        row=row_number,
                        field="description",
                        code="INVALID_DESCRIPTION",
                        message="Risk Description must be at least 10 characters",
                    )
                )

            category = self._normalise_category(category_raw)
            if not category:
                row_errors.append(
                    RowError(row=row_number, field="category", code="REQUIRED", message="Category is required")
                )

            inherent_impact = self._parse_score(
                row.get("inherent_impact"), field="inherent_impact", row=row_number, errors=row_errors
            )
            inherent_likelihood = self._parse_score(
                row.get("inherent_likelihood"),
                field="inherent_likelihood",
                row=row_number,
                errors=row_errors,
            )
            residual_impact = self._parse_score(
                row.get("residual_impact"), field="residual_impact", row=row_number, errors=row_errors
            )
            residual_likelihood = self._parse_score(
                row.get("residual_likelihood"),
                field="residual_likelihood",
                row=row_number,
                errors=row_errors,
            )

            if row_errors:
                errors.extend(row_errors)
                continue

            assert category is not None
            assert inherent_impact is not None
            assert inherent_likelihood is not None
            assert residual_impact is not None
            assert residual_likelihood is not None

            existing_risk = existing.get(reference)
            action = "update" if existing_risk else "create"
            if action == "create":
                creates += 1
            else:
                updates += 1

            validated.append(
                ValidatedImportRow(
                    row=row_number,
                    action=action,
                    reference=reference,
                    title=title,
                    description=description,
                    category=category,
                    inherent_likelihood=inherent_likelihood,
                    inherent_impact=inherent_impact,
                    residual_likelihood=residual_likelihood,
                    residual_impact=residual_impact,
                    risk_owner_name=owner,
                    treatment_plan=treatment_plan,
                    status=self._normalise_status(status_raw),
                    identified_date=self._parse_datetime(row.get("identified_date")),
                    last_review_date=self._parse_datetime(row.get("last_review_date")),
                    next_review_date=self._parse_datetime(row.get("next_review_date")),
                    review_notes=self._build_review_notes(comments, control_effectiveness),
                    context=root_causes or None,
                    existing_id=existing_risk.id if existing_risk else None,
                )
            )

        error_row_nums = {e.row for e in errors}
        preview = [
            {
                "row": item.row,
                "action": item.action,
                "reference": item.reference,
                "title": item.title,
                "category": item.category,
                "inherent_score": RiskScoringEngine.calculate_score(item.inherent_likelihood, item.inherent_impact),
                "residual_score": RiskScoringEngine.calculate_score(item.residual_likelihood, item.residual_impact),
                "risk_owner_name": item.risk_owner_name,
                "status": item.status,
            }
            for item in validated[:50]
        ]
        report = ImportValidationReport(
            dry_run=dry_run,
            total_rows=len(rows),
            valid_rows=len(validated),
            error_rows=len(error_row_nums),
            creates=creates,
            updates=updates,
            errors=errors,
            preview=preview,
        )
        return report, validated

    async def validate_action_plan_rows(
        self,
        rows: list[dict[str, Any]],
        *,
        tenant_id: int,
        register_refs: set[str],
    ) -> tuple[list[ValidatedActionPlanRow], list[RowError], int, int]:
        """Validate Action Plan rows; resolve CAPA create vs update via MatchKey/Action ID."""
        source_refs = {
            self._cell_text(row.get("match_key")) or self._cell_text(row.get("action_id"))
            for row in rows
            if self._cell_text(row.get("match_key")) or self._cell_text(row.get("action_id"))
        }
        existing_capas = await self._existing_capa_by_source_refs(source_refs, tenant_id)
        # Also resolve risks that already exist in DB (for dry-run when not creating in same file).
        linked_refs = {
            self._cell_text(row.get("risk_reference")).upper()
            for row in rows
            if self._cell_text(row.get("risk_reference"))
        }
        existing_risks = await self._existing_references(linked_refs, tenant_id)

        errors: list[RowError] = []
        validated: list[ValidatedActionPlanRow] = []
        seen_keys: dict[str, int] = {}
        creates = 0
        updates = 0

        for row in rows:
            row_number = int(row["__row__"])
            row_errors: list[RowError] = []
            action_id = self._cell_text(row.get("action_id"))
            risk_ref = self._cell_text(row.get("risk_reference")).upper()
            description = self._cell_text(row.get("description"))
            owner = self._cell_text(row.get("owner")) or None
            status_raw = self._cell_text(row.get("status")) or "open"
            match_key = self._cell_text(row.get("match_key")) or action_id
            progress_notes = self._cell_text(row.get("progress_notes")) or None

            if not action_id:
                row_errors.append(
                    RowError(row=row_number, field="action_id", code="REQUIRED", message="Action ID is required")
                )
            if not risk_ref:
                row_errors.append(
                    RowError(
                        row=row_number,
                        field="risk_reference",
                        code="REQUIRED",
                        message="Linked Risk Ref is required",
                    )
                )
            elif not REFERENCE_PATTERN.match(risk_ref):
                row_errors.append(
                    RowError(
                        row=row_number,
                        field="risk_reference",
                        code="INVALID_REFERENCE",
                        message=f"Linked Risk Ref '{risk_ref}' must match PELR* pattern",
                    )
                )
            elif risk_ref not in register_refs and risk_ref not in existing_risks:
                row_errors.append(
                    RowError(
                        row=row_number,
                        field="risk_reference",
                        code="UNKNOWN_RISK_REF",
                        message=(
                            f"Linked Risk Ref '{risk_ref}' not found in Risk Register sheet " "or existing register"
                        ),
                    )
                )
            if not description or len(description) < 5:
                row_errors.append(
                    RowError(
                        row=row_number,
                        field="description",
                        code="INVALID_DESCRIPTION",
                        message="Action Description must be at least 5 characters",
                    )
                )
            if match_key and match_key in seen_keys:
                row_errors.append(
                    RowError(
                        row=row_number,
                        field="match_key",
                        code="DUPLICATE_IN_FILE",
                        message=f"Duplicate MatchKey/Action ID '{match_key}' (also on row {seen_keys[match_key]})",
                    )
                )
            else:
                if match_key:
                    seen_keys[match_key] = row_number

            if row_errors:
                errors.extend(row_errors)
                continue

            existing_capa = existing_capas.get(match_key)
            action = "update" if existing_capa else "create"
            if action == "create":
                creates += 1
            else:
                updates += 1

            title = self._action_title(description, action_id)
            full_description = description
            if progress_notes:
                full_description = f"{description}\n\nProgress notes: {progress_notes}"
            if owner:
                full_description = f"{full_description}\n\nOwner (import): {owner}"

            validated.append(
                ValidatedActionPlanRow(
                    row=row_number,
                    action=action,
                    action_id=action_id,
                    risk_reference=risk_ref,
                    title=title,
                    description=full_description,
                    status=status_raw,
                    match_key=match_key,
                    owner=owner,
                    due_date=self._parse_datetime(row.get("deadline")),
                    progress_notes=progress_notes,
                    existing_capa_id=existing_capa.id if existing_capa else None,
                )
            )

        return validated, errors, creates, updates

    async def dry_run(self, content: bytes, *, tenant_id: int) -> ImportValidationReport:
        register_rows, action_rows = self.parse_workbook(content)
        report, validated = await self.validate_rows(register_rows, tenant_id=tenant_id, dry_run=True)
        register_refs = {item.reference for item in validated} | {
            self._cell_text(r.get("reference")).upper() for r in register_rows if self._cell_text(r.get("reference"))
        }

        if action_rows is None:
            report.action_plan_skipped = True
            return report

        report.action_plan_skipped = False
        ap_validated, ap_errors, ap_creates, ap_updates = await self.validate_action_plan_rows(
            action_rows,
            tenant_id=tenant_id,
            register_refs=register_refs,
        )
        report.action_plan_total_rows = len(action_rows)
        report.action_plan_creates = ap_creates
        report.action_plan_updates = ap_updates
        report.action_plan_errors = ap_errors
        report.action_plan_error_rows = len({e.row for e in ap_errors})
        report.action_plan_preview = [
            {
                "row": item.row,
                "action": item.action,
                "action_id": item.action_id,
                "risk_reference": item.risk_reference,
                "title": item.title,
                "status": item.status,
                "match_key": item.match_key,
            }
            for item in ap_validated[:50]
        ]
        return report

    async def _apply_row(
        self,
        item: ValidatedImportRow,
        *,
        tenant_id: int,
        user_id: int,
    ) -> tuple[EnterpriseRisk, str]:
        inherent_score = RiskScoringEngine.calculate_score(item.inherent_likelihood, item.inherent_impact)
        residual_score = RiskScoringEngine.calculate_score(item.residual_likelihood, item.residual_impact)
        appetite_threshold = await self._appetite_threshold(item.category)
        now = datetime.now(timezone.utc).replace(tzinfo=None)

        if item.action == "update" and item.existing_id is not None:
            result = await self.db.execute(
                select(EnterpriseRisk).where(
                    EnterpriseRisk.id == item.existing_id,
                    EnterpriseRisk.tenant_id == tenant_id,
                )
            )
            risk = result.scalar_one()
            risk.title = item.title
            risk.description = item.description
            risk.category = item.category
            risk.context = item.context
            risk.inherent_likelihood = item.inherent_likelihood
            risk.inherent_impact = item.inherent_impact
            risk.inherent_score = inherent_score
            risk.residual_likelihood = item.residual_likelihood
            risk.residual_impact = item.residual_impact
            risk.residual_score = residual_score
            risk.appetite_threshold = appetite_threshold
            risk.is_within_appetite = residual_score <= appetite_threshold
            risk.risk_owner_name = item.risk_owner_name
            risk.treatment_plan = item.treatment_plan
            risk.status = item.status
            risk.review_notes = item.review_notes
            if item.identified_date:
                risk.identified_date = item.identified_date
            if item.last_review_date:
                risk.last_review_date = item.last_review_date
            if item.next_review_date:
                risk.next_review_date = item.next_review_date
            risk.updated_at = now
            return risk, "update"

        review_frequency_days = 90
        next_review = item.next_review_date or (now + timedelta(days=review_frequency_days))
        risk = EnterpriseRisk(
            tenant_id=tenant_id,
            reference=item.reference,
            title=item.title,
            description=item.description,
            category=item.category,
            context=item.context,
            inherent_likelihood=item.inherent_likelihood,
            inherent_impact=item.inherent_impact,
            inherent_score=inherent_score,
            residual_likelihood=item.residual_likelihood,
            residual_impact=item.residual_impact,
            residual_score=residual_score,
            appetite_threshold=appetite_threshold,
            is_within_appetite=residual_score <= appetite_threshold,
            risk_owner_name=item.risk_owner_name,
            treatment_plan=item.treatment_plan,
            status=item.status,
            review_notes=item.review_notes,
            review_frequency_days=review_frequency_days,
            identified_date=item.identified_date or now,
            last_review_date=item.last_review_date,
            next_review_date=next_review,
            created_by=user_id,
        )
        self.db.add(risk)
        return risk, "create"

    async def _apply_action_plan_row(
        self,
        item: ValidatedActionPlanRow,
        *,
        risk_by_ref: dict[str, EnterpriseRisk],
        user_id: int,
        tenant_id: int,
    ) -> tuple[CAPAAction, str]:
        risk = risk_by_ref.get(item.risk_reference)
        if risk is None:
            raise ValidationError(
                f"Action Plan row {item.row}: risk {item.risk_reference} not resolved after register import",
                code="RISK_REGISTER_IMPORT_ACTION_PLAN_UNRESOLVED",
            )

        risk_service = RiskService(self.db)

        if item.action == "update" and item.existing_capa_id is not None:
            result = await self.db.execute(
                select(CAPAAction).where(
                    CAPAAction.id == item.existing_capa_id,
                    CAPAAction.tenant_id == tenant_id,
                )
            )
            capa = result.scalar_one()
            capa.title = item.title[:255]
            capa.description = item.description
            capa.source_id = risk.id
            capa.source_type = CAPASource.RISK
            capa.source_reference = item.match_key
            capa.due_date = item.due_date
            # Status normalisation via create helper path values
            status_key = (item.status or "open").strip().lower().replace(" ", "_")
            from src.domain.models.capa import CAPAStatus

            if status_key in {"completed", "closed", "done"}:
                capa.status = CAPAStatus.CLOSED
            elif status_key in {"in_progress", "in-progress", "progress"}:
                capa.status = CAPAStatus.IN_PROGRESS
            elif status_key in {"verification", "verifying"}:
                capa.status = CAPAStatus.VERIFICATION
            else:
                capa.status = CAPAStatus.OPEN
            await self.db.flush()
            return capa, "update"

        capa = await risk_service.create_capa_action_for_risk(
            risk,
            title=item.title,
            description=item.description,
            created_by_id=user_id,
            due_date=item.due_date,
            status=item.status,
            source_reference=item.match_key,
            commit=False,
        )
        return capa, "create"

    async def commit(
        self,
        content: bytes,
        *,
        user_id: int,
        tenant_id: int,
    ) -> ImportCommitResult:
        register_rows, action_rows = self.parse_workbook(content)
        report, validated = await self.validate_rows(register_rows, tenant_id=tenant_id, dry_run=False)
        register_refs = {item.reference for item in validated}

        ap_validated: list[ValidatedActionPlanRow] = []
        if action_rows is None:
            report.action_plan_skipped = True
        else:
            report.action_plan_skipped = False
            ap_validated, ap_errors, ap_creates, ap_updates = await self.validate_action_plan_rows(
                action_rows,
                tenant_id=tenant_id,
                register_refs=register_refs,
            )
            report.action_plan_total_rows = len(action_rows)
            report.action_plan_creates = ap_creates
            report.action_plan_updates = ap_updates
            report.action_plan_errors = ap_errors
            report.action_plan_error_rows = len({e.row for e in ap_errors})
            report.action_plan_preview = [
                {
                    "row": item.row,
                    "action": item.action,
                    "action_id": item.action_id,
                    "risk_reference": item.risk_reference,
                    "title": item.title,
                    "status": item.status,
                    "match_key": item.match_key,
                }
                for item in ap_validated[:50]
            ]

        if not report.ok:
            raise ValidationError(
                "Risk register import validation failed; fix row errors before commit",
                code="RISK_REGISTER_IMPORT_VALIDATION_FAILED",
                details=report.to_dict(),
            )

        created_ids: list[int] = []
        updated_ids: list[int] = []
        risk_by_ref: dict[str, EnterpriseRisk] = {}
        for item in validated:
            risk, action = await self._apply_row(item, tenant_id=tenant_id, user_id=user_id)
            await self.db.flush()
            risk_by_ref[item.reference] = risk
            if action == "create":
                created_ids.append(risk.id)
            else:
                updated_ids.append(risk.id)

        # Ensure Action Plan can resolve risks that already existed but weren't in this file batch.
        missing_refs = {ap.risk_reference for ap in ap_validated} - set(risk_by_ref)
        if missing_refs:
            existing = await self._existing_references(missing_refs, tenant_id)
            risk_by_ref.update(existing)

        capa_created_ids: list[int] = []
        capa_updated_count = 0
        for ap_item in ap_validated:
            capa, action = await self._apply_action_plan_row(
                ap_item,
                risk_by_ref=risk_by_ref,
                user_id=user_id,
                tenant_id=tenant_id,
            )
            if action == "create":
                capa_created_ids.append(capa.id)
            else:
                capa_updated_count += 1

        await self.db.commit()

        final_report = ImportValidationReport(
            dry_run=False,
            total_rows=report.total_rows,
            valid_rows=report.valid_rows,
            error_rows=0,
            creates=len(created_ids),
            updates=len(updated_ids),
            errors=[],
            preview=report.preview,
            action_plan_skipped=report.action_plan_skipped,
            action_plan_total_rows=report.action_plan_total_rows,
            action_plan_creates=len(capa_created_ids),
            action_plan_updates=capa_updated_count,
            action_plan_error_rows=0,
            action_plan_errors=[],
            action_plan_preview=report.action_plan_preview,
        )
        return ImportCommitResult(
            created_count=len(created_ids),
            updated_count=len(updated_ids),
            created_risk_ids=created_ids,
            updated_risk_ids=updated_ids,
            capa_created_count=len(capa_created_ids),
            capa_updated_count=capa_updated_count,
            capa_created_ids=capa_created_ids,
            report=final_report,
        )

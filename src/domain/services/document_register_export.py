"""IMS052 Master Document Register export (WA-3 / L-07).

Enhances Export Center's existing ``documents`` module — one SoT, no twin
register export. Column contract is fixed: the builder never accepts a
column picker / visible-columns argument. Legacy IMS / PLA headers stay
present and empty until a future backfill lands (no fields exist today).
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import Select, select
from sqlalchemy.ext.asyncio import AsyncSession

from src.domain.models.document import Document
from src.domain.models.document_library import DocumentCategory, DocumentFunction
from src.domain.models.location import Location
from src.domain.services.document_library_rbac import user_can_read_library_document
from src.domain.services.href_registry import document_href

# Fixed IMS052 evidence-pack header — order is contract. Do not derive from UI.
IMS052_COLUMNS: tuple[str, ...] = (
    "PEL Reference",
    "Legacy IMS Ref",
    "Legacy PLA Ref",
    "Document Name",
    "Reference",
    "Issue",
    "Status",
    "Function",
    "Category",
    "Last Review Date",
    "Next Review Date",
    "Location",
    "Access Rights",
    "Retention",
    "Hyperlink",
)

IMS052_SHEET_TITLE = "IMS052 Register"
IMS052_FILENAME_STEM = "ims052_document_register"


def _enum_str(value: Any) -> str:
    if value is None:
        return ""
    return str(getattr(value, "value", value))


def _date_only(value: datetime | date | None) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value.isoformat()


def _excel_safe(value: str) -> str:
    """Prefix formula-like cells so free-text titles cannot inject Excel formulas."""
    if value and value[0] in ("=", "+", "-", "@"):
        return f"'{value}"
    return value


def build_register_row(
    doc: Any,
    *,
    function_name: str = "",
    category_name: str = "",
    location_name: str = "",
) -> list[str]:
    """Map one Document (+ joined labels) onto the fixed IMS052 column order."""
    cells = [
        str(getattr(doc, "pel_doc_ref", None) or ""),
        "",  # Legacy IMS — no field yet
        "",  # Legacy PLA — no field yet
        str(getattr(doc, "title", None) or ""),
        str(getattr(doc, "reference_number", None) or ""),
        str(getattr(doc, "version", None) or ""),
        _enum_str(getattr(doc, "status", None)),
        function_name,
        category_name,
        _date_only(getattr(doc, "reviewed_at", None)),
        _date_only(getattr(doc, "review_date", None)),
        location_name,
        str(getattr(doc, "access_level", None) or ""),
        _date_only(getattr(doc, "retention_until", None)),
        document_href(int(doc.id)),
    ]
    return [_excel_safe(cell) for cell in cells]


def rows_to_csv(rows: Sequence[Sequence[str]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(list(IMS052_COLUMNS))
    for row in rows:
        writer.writerow(list(row))
    return buffer.getvalue().encode("utf-8")


def rows_to_xlsx(rows: Sequence[Sequence[str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = IMS052_SHEET_TITLE
    sheet.append(list(IMS052_COLUMNS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append(list(row))
    sheet.freeze_panes = "A2"
    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


def rows_to_pdf(rows: Sequence[Sequence[str]]) -> bytes:
    """Minimal landscape evidence PDF — same rows as CSV/XLSX, clipped for width."""
    from fpdf import FPDF

    # Narrow printable subset for landscape A4; full pack remains in XLSX/CSV.
    pdf_cols = (
        "PEL Reference",
        "Document Name",
        "Reference",
        "Issue",
        "Status",
        "Next Review Date",
        "Access Rights",
    )
    indexes = [IMS052_COLUMNS.index(name) for name in pdf_cols]

    pdf = FPDF(orientation="L", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "IMS052 Master Document Register", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 6, "Fixed column contract - UI column picker ignored", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    col_widths = [32, 70, 32, 16, 22, 28, 28]
    pdf.set_font("Helvetica", "B", 7)
    for width, name in zip(col_widths, pdf_cols):
        pdf.cell(width, 6, name[:28], border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 7)
    for row in rows:
        for width, idx in zip(col_widths, indexes):
            text = str(row[idx] if idx < len(row) else "")[:40]
            pdf.cell(width, 5, text, border=1)
        pdf.ln()

    return bytes(pdf.output())


async def build_document_register_rows(
    db: AsyncSession,
    tenant_id: int,
    *,
    user: Any,
    row_limit: int,
) -> tuple[list[list[str]], int, bool]:
    """Load active Register estate rows the caller may see (matches list ACL).

    Returns ``(data_rows, total_visible, truncated)``. ``total_visible`` is
    post-ACL so the count cannot leak restricted rows.
    """
    stmt: Select[Any] = (
        select(Document)
        .where(Document.tenant_id == tenant_id)
        .where(Document.is_active.is_(True))
        .order_by(Document.id.desc())
    )
    result = await db.execute(stmt)
    documents = list(result.scalars().all())

    # Taxonomy map for restricted ACL (same batch pattern as list_documents).
    restricted_cat_ids = {
        d.category_id
        for d in documents
        if d.category_id is not None and (getattr(d, "access_level", None) or "") == "restricted"
    }
    taxonomy_by_cat: dict[int, str] = {}
    if restricted_cat_ids:
        cat_rows = await db.execute(
            select(DocumentCategory.id, DocumentCategory.taxonomy_id).where(DocumentCategory.id.in_(restricted_cat_ids))
        )
        taxonomy_by_cat = {row[0]: row[1] for row in cat_rows.all()}

    visible: list[Document] = []
    for doc in documents:
        tax = taxonomy_by_cat.get(doc.category_id) if doc.category_id else None
        if user_can_read_library_document(doc, user, taxonomy_id=tax):
            visible.append(doc)

    # ACL must run across the whole active estate before the export cap is
    # applied. Limiting the SQL query first can omit older readable documents,
    # and an active-row count cannot produce an exact post-ACL total.
    total_visible = len(visible)
    truncated = total_visible > row_limit
    if truncated:
        visible = visible[:row_limit]

    function_ids = {d.function_id for d in visible if d.function_id is not None}
    category_ids = {d.category_id for d in visible if d.category_id is not None}
    location_ids = {d.site_location_id for d in visible if d.site_location_id is not None}

    function_names: dict[int, str] = {}
    if function_ids:
        rows = await db.execute(
            select(DocumentFunction.id, DocumentFunction.name).where(DocumentFunction.id.in_(function_ids))
        )
        function_names = {row[0]: row[1] for row in rows.all()}

    category_names: dict[int, str] = {}
    if category_ids:
        rows = await db.execute(
            select(DocumentCategory.id, DocumentCategory.name).where(DocumentCategory.id.in_(category_ids))
        )
        category_names = {row[0]: row[1] for row in rows.all()}

    location_names: dict[int, str] = {}
    if location_ids:
        rows = await db.execute(select(Location.id, Location.name).where(Location.id.in_(location_ids)))
        location_names = {row[0]: row[1] for row in rows.all()}

    data_rows = [
        build_register_row(
            doc,
            function_name=function_names.get(doc.function_id, "") if doc.function_id else "",
            category_name=category_names.get(doc.category_id, "") if doc.category_id else "",
            location_name=location_names.get(doc.site_location_id, "") if doc.site_location_id else "",
        )
        for doc in visible
    ]
    return data_rows, total_visible, truncated


def serialize_register(
    rows: Sequence[Sequence[str]],
    export_format: str,
) -> tuple[bytes, str, str]:
    """Return ``(content, media_type, extension)`` for the fixed IMS052 rows."""
    fmt = (export_format or "csv").strip().lower()
    if fmt == "csv":
        return rows_to_csv(rows), "text/csv; charset=utf-8", "csv"
    if fmt == "xlsx":
        return (
            rows_to_xlsx(rows),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )
    if fmt == "pdf":
        return rows_to_pdf(rows), "application/pdf", "pdf"
    raise ValueError(f"Unsupported register export format '{export_format}'")

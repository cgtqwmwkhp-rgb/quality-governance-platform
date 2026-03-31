"""Shared document extraction service for reusable native text extraction."""

from __future__ import annotations

import csv
import io
import logging
import zipfile
from dataclasses import dataclass, field

from defusedxml import ElementTree as ET

from src.domain.models.document import FileType

logger = logging.getLogger(__name__)


_SYMBOL_MAP = {
    "\u2713": "[PASS]",  # ✓
    "\u2714": "[PASS]",  # ✔
    "\u2611": "[PASS]",  # ☑
    "\u2705": "[PASS]",  # ✅
    "\u2717": "[FAIL]",  # ✗
    "\u2718": "[FAIL]",  # ✘
    "\u2612": "[FAIL]",  # ☒
    "\u274c": "[FAIL]",  # ❌
    "\u2610": "[UNCHECKED]",  # ☐
    "\u2605": "[STAR_FILLED]",  # ★
    "\u2606": "[STAR_EMPTY]",  # ☆
    "\u25cf": "[FILLED]",  # ●
    "\u25cb": "[EMPTY]",  # ○
}


def _normalize_symbols(text: str) -> str:
    """Replace Unicode checkmarks, crosses, and stars with semantic tags."""
    for symbol, tag in _SYMBOL_MAP.items():
        if symbol in text:
            text = text.replace(symbol, f" {tag} ")
    return text


@dataclass
class ExtractedDocumentContent:
    """Normalized extraction result shared across document and audit workflows."""

    text: str
    page_count: int | None = None
    sheet_count: int | None = None
    has_tables: bool = False
    note: str | None = None
    page_texts: list[str] = field(default_factory=list)
    extraction_method: str = "native"
    color_annotations: list[dict[str, object]] = field(default_factory=list)


def extract_document_content(file_type: FileType, file_name: str, content: bytes) -> ExtractedDocumentContent:
    """Return searchable text for supported document formats."""
    if file_type in {FileType.TXT, FileType.MD}:
        return ExtractedDocumentContent(text=content.decode("utf-8", errors="ignore"))

    if file_type == FileType.CSV:
        decoded = content.decode("utf-8", errors="ignore")
        rows = [", ".join(cell.strip() for cell in row if cell.strip()) for row in csv.reader(io.StringIO(decoded))]
        filtered_rows = [row for row in rows if row]
        return ExtractedDocumentContent(text="\n".join(filtered_rows), has_tables=bool(filtered_rows))

    if file_type == FileType.PDF:
        return _extract_pdf_text(content, file_name)

    if file_type == FileType.DOCX:
        return _extract_docx_text(content, file_name)

    if file_type == FileType.XLSX:
        return _extract_xlsx_text(content, file_name)

    unsupported_notes = {
        FileType.DOC: "Stored successfully but legacy .doc extraction is not supported yet.",
        FileType.XLS: "Stored successfully but legacy .xls extraction is not supported yet.",
        FileType.PNG: "Stored successfully but image OCR is not supported yet.",
        FileType.JPG: "Stored successfully but image OCR is not supported yet.",
        FileType.JPEG: "Stored successfully but image OCR is not supported yet.",
    }
    return ExtractedDocumentContent(text="", note=unsupported_notes.get(file_type))


def _extract_pdf_text(content: bytes, file_name: str) -> ExtractedDocumentContent:
    """Extract searchable text from PDF using pdfplumber (table-aware) with pypdf fallback."""
    plumber_result = _extract_pdf_via_pdfplumber(content, file_name)
    if plumber_result is not None:
        return plumber_result

    try:
        from pypdf import PdfReader
    except ImportError:
        logger.warning("PDF extraction dependency missing for %s", file_name)
        return ExtractedDocumentContent(
            text="",
            note="Stored successfully but PDF extraction is unavailable in this environment.",
        )

    try:
        reader = PdfReader(io.BytesIO(content))
        page_text = [_normalize_symbols((page.extract_text() or "").strip()) for page in reader.pages]
        filtered_pages = [part for part in page_text if part]
        text = "\n\n".join(filtered_pages)
        return ExtractedDocumentContent(
            text=text,
            page_count=len(reader.pages),
            page_texts=page_text,
        )
    except Exception as exc:
        logger.warning("PDF extraction failed for %s: %s", file_name, type(exc).__name__)
        return ExtractedDocumentContent(
            text="",
            note=f"Stored successfully but PDF extraction failed for {file_name}.",
        )


def _classify_color(color) -> str | None:
    """Map an RGB color tuple to a semantic label (pass/fail/observation)."""
    if not color or not isinstance(color, (list, tuple)) or len(color) < 3:
        return None
    try:
        r, g, b = float(color[0]), float(color[1]), float(color[2])
    except (TypeError, ValueError):
        return None
    if r <= 1 and g <= 1 and b <= 1:
        r, g, b = r * 255, g * 255, b * 255
    if g > 150 and r < 100 and b < 100:
        return "pass"
    if r > 150 and g < 100 and b < 100:
        return "fail"
    if r > 200 and g > 150 and b < 100:
        return "observation"
    if r > 200 and g > 200 and b < 100:
        return "observation"
    return None


def _extract_pdf_via_pdfplumber(content: bytes, file_name: str) -> ExtractedDocumentContent | None:
    """Use pdfplumber for table-aware PDF extraction with color detection."""
    try:
        import pdfplumber  # noqa: F811
    except ImportError:
        return None

    try:
        page_texts: list[str] = []
        has_tables = False
        color_annotations: list[dict[str, object]] = []

        with pdfplumber.open(io.BytesIO(content)) as pdf:
            page_count = len(pdf.pages)
            for page_num, page in enumerate(pdf.pages, start=1):
                parts: list[str] = []
                tables = page.extract_tables()
                if tables:
                    has_tables = True
                    for table in tables:
                        for row in table:
                            cells = [str(c).strip() for c in row if c]
                            if cells:
                                parts.append(" | ".join(cells))

                try:
                    for rect in page.rects or []:
                        fill = rect.get("non_stroking_color")
                        label = _classify_color(fill)
                        if label:
                            color_annotations.append(
                                {
                                    "page": page_num,
                                    "type": "cell_fill",
                                    "color_label": label,
                                    "x0": rect.get("x0"),
                                    "y0": rect.get("y0"),
                                }
                            )
                except Exception:
                    pass

                plain = page.extract_text()
                if plain:
                    parts.append(plain.strip())
                page_text = _normalize_symbols("\n".join(parts))
                page_texts.append(page_text)

        filtered = [p for p in page_texts if p.strip()]
        text = "\n\n".join(filtered)
        if not text.strip():
            return None

        return ExtractedDocumentContent(
            text=text,
            page_count=page_count,
            page_texts=page_texts,
            has_tables=has_tables,
            extraction_method="pdfplumber",
            color_annotations=color_annotations,
        )
    except Exception as exc:
        logger.warning("pdfplumber extraction failed for %s: %s — falling back to pypdf", file_name, type(exc).__name__)
        return None


def _extract_docx_via_python_docx(content: bytes, file_name: str) -> ExtractedDocumentContent | None:
    """Extract text and tables from DOCX using python-docx. Returns None if unavailable."""
    try:
        from docx import Document
    except ImportError:
        return None

    try:
        doc = Document(io.BytesIO(content))
        parts: list[str] = []
        has_tables = False

        for element in doc.element.body:
            tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag
            if tag == "p":
                text = element.text or ""
                runs = element.findall(".//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t")
                if runs:
                    text = "".join(r.text or "" for r in runs)
                text = text.strip()
                if text:
                    parts.append(text)
            elif tag == "tbl":
                has_tables = True

        for table in doc.tables:
            for row in table.rows:
                cells = [cell.text.strip() for cell in row.cells if cell.text.strip()]
                if cells:
                    parts.append(" | ".join(cells))

        text = _normalize_symbols("\n".join(parts))
        return ExtractedDocumentContent(
            text=text,
            has_tables=has_tables,
            extraction_method="python_docx",
        )
    except Exception as exc:
        logger.warning("python-docx extraction failed for %s: %s", file_name, type(exc).__name__)
        return None


def _extract_docx_text(content: bytes, file_name: str) -> ExtractedDocumentContent:
    """Extract text from DOCX, preferring python-docx for table support."""
    docx_result = _extract_docx_via_python_docx(content, file_name)
    if docx_result is not None:
        return docx_result

    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            document_xml = archive.read("word/document.xml")
    except Exception as exc:
        logger.warning("DOCX extraction failed for %s: %s", file_name, type(exc).__name__)
        return ExtractedDocumentContent(
            text="",
            note=f"Stored successfully but DOCX extraction failed for {file_name}.",
        )

    root = ET.fromstring(document_xml)
    paragraphs: list[str] = []
    for paragraph in root.iter():
        if paragraph.tag.endswith("}p"):
            runs = [node.text or "" for node in paragraph.iter() if node.tag.endswith("}t") and node.text]
            joined = "".join(runs).strip()
            if joined:
                paragraphs.append(joined)
    return ExtractedDocumentContent(text=_normalize_symbols("\n".join(paragraphs)))


def _read_xlsx_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    """Load workbook shared strings for cell lookup."""
    try:
        raw = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []

    root = ET.fromstring(raw)
    values: list[str] = []
    for string_item in root.iter():
        if string_item.tag.endswith("}si"):
            parts = [node.text or "" for node in string_item.iter() if node.tag.endswith("}t") and node.text]
            values.append("".join(parts))
    return values


def _extract_xlsx_via_openpyxl(content: bytes, file_name: str) -> ExtractedDocumentContent | None:
    """Extract text and cell colors from XLSX using openpyxl. Returns None if unavailable."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    try:
        wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
        sheet_lines: list[str] = []
        color_annotations: list[dict[str, object]] = []

        for idx, ws in enumerate(wb.worksheets, start=1):
            rows: list[str] = []
            for row in ws.iter_rows():
                values: list[str] = []
                for cell in row:
                    val = str(cell.value).strip() if cell.value is not None else ""
                    if val:
                        fill = cell.fill
                        if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
                            rgb_hex = str(fill.fgColor.rgb)
                            color_label = _classify_hex_color(rgb_hex)
                            if color_label:
                                val = f"{val} [{color_label.upper()}]"
                                color_annotations.append(
                                    {
                                        "sheet": idx,
                                        "cell": cell.coordinate,
                                        "color_hex": rgb_hex,
                                        "color_label": color_label,
                                        "value": str(cell.value),
                                    }
                                )
                        values.append(val)
                if values:
                    rows.append(", ".join(values))
            if rows:
                sheet_lines.append(f"Sheet {idx}")
                sheet_lines.extend(rows)

        wb.close()
        text = _normalize_symbols("\n".join(sheet_lines))
        return ExtractedDocumentContent(
            text=text,
            sheet_count=len(wb.sheetnames),
            has_tables=bool(sheet_lines),
            extraction_method="openpyxl",
            color_annotations=color_annotations,
        )
    except Exception as exc:
        logger.warning("openpyxl extraction failed for %s: %s", file_name, type(exc).__name__)
        return None


def _classify_hex_color(rgb_hex: str) -> str | None:
    """Classify an ARGB/RGB hex string to a semantic label."""
    hex_str = rgb_hex.lstrip("#")
    if len(hex_str) == 8:
        hex_str = hex_str[2:]
    if len(hex_str) != 6:
        return None
    try:
        r = int(hex_str[0:2], 16)
        g = int(hex_str[2:4], 16)
        b = int(hex_str[4:6], 16)
    except ValueError:
        return None
    return _classify_color((r / 255, g / 255, b / 255))


def _extract_xlsx_text(content: bytes, file_name: str) -> ExtractedDocumentContent:
    """Extract text rows from XLSX, preferring openpyxl for color support."""
    openpyxl_result = _extract_xlsx_via_openpyxl(content, file_name)
    if openpyxl_result is not None:
        return openpyxl_result

    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            shared_strings = _read_xlsx_shared_strings(archive)
            sheet_paths = sorted(name for name in archive.namelist() if name.startswith("xl/worksheets/sheet"))
            sheet_lines: list[str] = []

            for idx, sheet_path in enumerate(sheet_paths, start=1):
                root = ET.fromstring(archive.read(sheet_path))
                rows: list[str] = []
                for row in root.iter():
                    if not row.tag.endswith("}row"):
                        continue
                    values: list[str] = []
                    for cell in row:
                        if not cell.tag.endswith("}c"):
                            continue
                        cell_type = cell.attrib.get("t")
                        value = ""
                        inline_value = next(
                            (node.text for node in cell.iter() if node.tag.endswith("}t") and node.text),
                            None,
                        )
                        if inline_value:
                            value = inline_value
                        else:
                            raw_value = next(
                                (node.text for node in cell.iter() if node.tag.endswith("}v") and node.text),
                                "",
                            )
                            if cell_type == "s" and raw_value.isdigit():
                                shared_index = int(raw_value)
                                value = (
                                    shared_strings[shared_index] if shared_index < len(shared_strings) else raw_value
                                )
                            else:
                                value = raw_value
                        value = value.strip()
                        if value:
                            values.append(value)
                    if values:
                        rows.append(", ".join(values))

                if rows:
                    sheet_lines.append(f"Sheet {idx}")
                    sheet_lines.extend(rows)

        return ExtractedDocumentContent(
            text=_normalize_symbols("\n".join(sheet_lines)),
            sheet_count=len(sheet_paths),
            has_tables=bool(sheet_lines),
        )
    except Exception as exc:
        logger.warning("XLSX extraction failed for %s: %s", file_name, type(exc).__name__)
        return ExtractedDocumentContent(
            text="",
            note=f"Stored successfully but XLSX extraction failed for {file_name}.",
        )

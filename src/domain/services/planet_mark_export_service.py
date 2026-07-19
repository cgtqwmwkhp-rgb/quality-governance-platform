"""Assemble Planet Mark year export packs (JSON + XLSX) from live domain data."""

from __future__ import annotations

import io
import json
import math
import re
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.domain.models.planet_mark import CarbonReportingYear, ImprovementAction, Scope3CategoryData


def _has_positive(value: float | None) -> bool:
    return value is not None and math.isfinite(value) and value > 0


def build_hotspot_initiatives(
    categories: list[dict[str, Any]],
    total_co2e: float,
    *,
    limit: int = 5,
) -> list[dict[str, Any]]:
    """Rank measured Scope 3 categories into SMART-ish initiative drafts (FE parity)."""
    if not categories:
        return []

    total = total_co2e if _has_positive(total_co2e) else None
    ranked: list[tuple[dict[str, Any], float]] = []

    for category in categories:
        if not category.get("is_measured") or not _has_positive(category.get("total_co2e")):
            continue
        percentage = category.get("percentage")
        if _has_positive(percentage):
            footprint_percent = float(percentage)
        elif total:
            footprint_percent = (float(category["total_co2e"]) / total) * 100
        else:
            footprint_percent = 0.0
        ranked.append((category, footprint_percent))

    ranked.sort(key=lambda item: item[1], reverse=True)
    ranked = ranked[:limit]

    initiatives: list[dict[str, Any]] = []
    for category, footprint_percent in ranked:
        suggested_reduction_percent = 10 if footprint_percent >= 20 else 5
        current_co2e = float(category["total_co2e"])
        expected_reduction_co2e = (current_co2e * suggested_reduction_percent) / 100
        initiatives.append(
            {
                "id": f"cat-{category['number']}",
                "title": f"Reduce {category['name']} emissions by {suggested_reduction_percent}%",
                "categoryNumber": category["number"],
                "categoryName": category["name"],
                "footprintPercent": footprint_percent,
                "currentCo2e": current_co2e,
                "suggestedReductionPercent": suggested_reduction_percent,
                "expectedReductionCo2e": expected_reduction_co2e,
                "specific": (
                    f"Target Scope 3 category {category['number']} ({category['name']}) "
                    "hotspot ranked by footprint share."
                ),
                "measurable": (
                    f"Cut category emissions by ~{suggested_reduction_percent}% "
                    f"(~{expected_reduction_co2e:.2f} tCO₂e) vs current {current_co2e:.2f} tCO₂e."
                ),
            }
        )
    return initiatives


def export_pack_filename(year_label: str, ext: str) -> str:
    safe_label = re.sub(r"[^\w-]+", "_", year_label)
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return f"planet-mark-export-{safe_label}-{stamp}.{ext}"


class PlanetMarkExportService:
    """Build JSON/XLSX export packs aligned with FE ``buildPlanetMarkExportPack``."""

    def __init__(self, db: AsyncSession):
        self.db = db

    async def assemble_pack(self, *, year_id: int, tenant_id: int) -> dict[str, Any]:
        year = (
            await self.db.execute(
                select(CarbonReportingYear).where(
                    CarbonReportingYear.id == year_id,
                    CarbonReportingYear.tenant_id == tenant_id,
                )
            )
        ).scalar_one_or_none()
        if year is None:
            raise ValueError("Reporting year not found")

        scope3_rows: list[Scope3CategoryData] = list(
            (
                await self.db.execute(
                    select(Scope3CategoryData)
                    .where(
                        Scope3CategoryData.reporting_year_id == year_id,
                        Scope3CategoryData.tenant_id == tenant_id,
                    )
                    .order_by(Scope3CategoryData.category_number)
                )
            )
            .scalars()
            .all()
        )

        action_rows: list[ImprovementAction] = list(
            (
                await self.db.execute(
                    select(ImprovementAction)
                    .where(
                        ImprovementAction.reporting_year_id == year_id,
                        ImprovementAction.tenant_id == tenant_id,
                    )
                    .order_by(ImprovementAction.time_bound)
                )
            )
            .scalars()
            .all()
        )

        scope3_total = sum(c.total_co2e for c in scope3_rows)
        scope3_categories = [
            {
                "number": c.category_number,
                "name": c.category_name,
                "is_measured": c.is_measured,
                "total_co2e": c.total_co2e,
                "percentage": (round((c.total_co2e / scope3_total * 100), 1) if scope3_total > 0 else 0),
            }
            for c in scope3_rows
        ]

        now = datetime.now(timezone.utc).replace(tzinfo=None)
        improvement_actions = [
            {
                "id": action.id,
                "action_id": action.action_id,
                "action_title": action.action_title,
                "owner": action.achievable_owner,
                "deadline": action.time_bound.isoformat(),
                "status": action.status,
                "progress_percent": action.progress_percent,
                "expected_reduction_pct": action.expected_reduction_pct,
                "is_overdue": action.status != "completed" and action.time_bound < now,
            }
            for action in action_rows
        ]

        period = (
            f"{year.period_start.strftime('%d %b %Y')} - {year.period_end.strftime('%d %b %Y')}"
        )

        return {
            "export_kind": "json_pack",
            "pdf_note": (
                "Branded PDF pack is a follow-on — JSON and XLSX packs are live from the "
                "authenticated export API."
            ),
            "xlsx_note": "Use ?format=xlsx on the export endpoint for the spreadsheet pack.",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "reporting_year": {
                "id": year.id,
                "year_label": year.year_label,
                "year_number": year.year_number,
                "period": period,
                "average_fte": year.average_fte,
                "total_emissions": year.total_emissions,
                "emissions_per_fte": year.emissions_per_fte,
                "scope_1": year.scope_1_total,
                "scope_2_market": year.scope_2_market,
                "scope_3": year.scope_3_total,
                "data_quality": year.overall_data_quality,
                "certification_status": year.certification_status,
                "is_baseline": year.is_baseline_year,
            },
            "scope3_categories": scope3_categories,
            "improvement_actions": improvement_actions,
            "hotspot_initiatives": build_hotspot_initiatives(scope3_categories, scope3_total),
        }

    def build_json_bytes(self, pack: dict[str, Any]) -> bytes:
        return json.dumps(pack, indent=2).encode("utf-8")

    def build_xlsx_bytes(self, pack: dict[str, Any]) -> bytes:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary.append(["Planet Mark export pack"])
        summary.append(["Generated at", pack.get("generated_at")])
        summary.append(["PDF note", pack.get("pdf_note")])
        summary.append([])

        year = pack.get("reporting_year") or {}
        summary.append(["Reporting year", year.get("year_label")])
        summary.append(["Period", year.get("period")])
        summary.append(["Average FTE", year.get("average_fte")])
        summary.append(["Total emissions (tCO₂e)", year.get("total_emissions")])
        summary.append(["Emissions per FTE", year.get("emissions_per_fte")])
        summary.append(["Scope 1", year.get("scope_1")])
        summary.append(["Scope 2 (market)", year.get("scope_2_market")])
        summary.append(["Scope 3", year.get("scope_3")])
        summary.append(["Data quality", year.get("data_quality")])
        summary.append(["Certification status", year.get("certification_status")])

        scope3_sheet = workbook.create_sheet("Scope 3")
        scope3_sheet.append(["Number", "Name", "Measured", "tCO₂e", "Percentage"])
        for row in pack.get("scope3_categories") or []:
            scope3_sheet.append(
                [
                    row.get("number"),
                    row.get("name"),
                    row.get("is_measured"),
                    row.get("total_co2e"),
                    row.get("percentage"),
                ]
            )

        actions_sheet = workbook.create_sheet("Actions")
        actions_sheet.append(
            [
                "ID",
                "Action ID",
                "Title",
                "Owner",
                "Deadline",
                "Status",
                "Progress %",
                "Expected reduction %",
                "Overdue",
            ]
        )
        for row in pack.get("improvement_actions") or []:
            actions_sheet.append(
                [
                    row.get("id"),
                    row.get("action_id"),
                    row.get("action_title"),
                    row.get("owner"),
                    row.get("deadline"),
                    row.get("status"),
                    row.get("progress_percent"),
                    row.get("expected_reduction_pct"),
                    row.get("is_overdue"),
                ]
            )

        initiatives_sheet = workbook.create_sheet("Hotspot initiatives")
        initiatives_sheet.append(
            [
                "ID",
                "Title",
                "Category #",
                "Category",
                "Footprint %",
                "Current tCO₂e",
                "Suggested reduction %",
                "Expected reduction tCO₂e",
            ]
        )
        for row in pack.get("hotspot_initiatives") or []:
            initiatives_sheet.append(
                [
                    row.get("id"),
                    row.get("title"),
                    row.get("categoryNumber"),
                    row.get("categoryName"),
                    row.get("footprintPercent"),
                    row.get("currentCo2e"),
                    row.get("suggestedReductionPercent"),
                    row.get("expectedReductionCo2e"),
                ]
            )

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

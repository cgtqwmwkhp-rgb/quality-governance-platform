"""Parse Atlas / Citation Training Matrix Report (CSV or .xlsx).

Two header layouts are accepted:

- Atlas CSV: title row, then course names, then Status/Passed/Expiry, then people.
- Citation Excel: ``Full Name`` / ``Department`` / course names on row 1,
  Status / Completed date / Expiry Date on row 2, people from row 3.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional


def normalize_course_key(name: str) -> str:
    key = re.sub(r"[^a-z0-9]+", "_", (name or "").strip().lower()).strip("_")
    return key[:255] or "course"


def normalize_person_name(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip())


def person_name_match_keys(name: str | None) -> set[str]:
    """Match Atlas 'David Harris' ↔ 'Harris, David' style variants (lowercased)."""
    normalized = normalize_person_name(name or "").lower()
    if not normalized:
        return set()
    keys = {normalized}
    if "," in normalized:
        last, first = [part.strip() for part in normalized.split(",", 1)]
        if first and last:
            keys.add(f"{first} {last}")
    else:
        parts = normalized.split()
        if len(parts) >= 2:
            keys.add(f"{parts[-1]}, {' '.join(parts[:-1])}")
    return keys


def parse_atlas_date(value: str | None) -> Optional[date]:
    raw = (value or "").strip()
    if not raw:
        return None
    if "T" in raw:
        raw = raw.split("T", 1)[0]
    if " " in raw:
        raw = raw.split(" ", 1)[0]
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _header_status(rows: list[list[str]], row_idx: int) -> str:
    if row_idx >= len(rows) or len(rows[row_idx]) < 3:
        return ""
    return (rows[row_idx][2] or "").strip().lower()


@dataclass
class ParsedCell:
    course_name: str
    course_key: str
    atlas_status: Optional[str]
    passed_on: Optional[date]
    expires_on: Optional[date]


@dataclass
class ParsedPerson:
    atlas_name: str
    department: Optional[str]
    cells: list[ParsedCell] = field(default_factory=list)


@dataclass
class ParsedMatrix:
    courses: list[str]
    people: list[ParsedPerson]
    cell_count: int
    nonempty_cell_count: int
    expiry_without_passed_count: int


def parse_training_matrix_rows(rows: list[list[str]]) -> ParsedMatrix:
    """Parse already-stringified matrix rows (CSV or Excel)."""
    if len(rows) < 3:
        raise ValueError("Training matrix must include course names, a Status header row, and people")

    status_r1 = _header_status(rows, 1)
    status_r2 = _header_status(rows, 2)
    if status_r1.startswith("status"):
        course_row_idx, people_start = 0, 2
    elif status_r2.startswith("status"):
        course_row_idx, people_start = 1, 3
    else:
        raise ValueError(
            "Training matrix must include course names and a Status / date header row. "
            "This is not an Atlas or Citation Training Matrix Report."
        )

    if people_start >= len(rows):
        raise ValueError("Training matrix must include people rows")

    course_row = rows[course_row_idx]
    courses: list[str] = []
    for i in range(2, len(course_row), 3):
        name = (course_row[i] or "").strip()
        if name:
            courses.append(name)

    if not courses:
        raise ValueError("No course columns found in training matrix")

    people_by_name: dict[str, ParsedPerson] = {}
    cell_count = 0
    nonempty = 0
    expiry_without_passed = 0

    for row in rows[people_start:]:
        if not row or not (row[0] or "").strip():
            continue
        # Ignore footer noise
        name = normalize_person_name(row[0])
        if name.lower().startswith("page "):
            continue
        department = (row[1] if len(row) > 1 else "").strip() or None
        person_key = name.lower()
        person = people_by_name.get(person_key)
        if person is None:
            person = ParsedPerson(atlas_name=name, department=department)
            people_by_name[person_key] = person
        elif department:
            person.department = department

        cells_by_course = {c.course_key: c for c in person.cells}
        for ci, course_name in enumerate(courses):
            base = 2 + ci * 3
            cell_count += 1
            status = (row[base] if base < len(row) else "").strip() or None
            passed_s = (row[base + 1] if base + 1 < len(row) else "").strip()
            expiry_s = (row[base + 2] if base + 2 < len(row) else "").strip()
            if not (status or passed_s or expiry_s):
                continue
            if status and status.lower().startswith("page "):
                continue
            nonempty += 1
            passed_on = parse_atlas_date(passed_s)
            expires_on = parse_atlas_date(expiry_s)
            if expires_on and not passed_on:
                expiry_without_passed += 1
            course_key = normalize_course_key(course_name)
            cells_by_course[course_key] = ParsedCell(
                course_name=course_name,
                course_key=course_key,
                atlas_status=status,
                passed_on=passed_on,
                expires_on=expires_on,
            )
        person.cells = list(cells_by_course.values())

    people = list(people_by_name.values())
    if not people:
        raise ValueError("No people rows found in training matrix")

    return ParsedMatrix(
        courses=courses,
        people=people,
        cell_count=cell_count,
        nonempty_cell_count=nonempty,
        expiry_without_passed_count=expiry_without_passed,
    )


def parse_training_matrix_csv(content: bytes | str) -> ParsedMatrix:
    if isinstance(content, bytes):
        text = content.decode("utf-8-sig", errors="replace")
    else:
        text = content
    rows = [list(row) for row in csv.reader(io.StringIO(text))]
    return parse_training_matrix_rows(rows)


def parse_training_matrix_xlsx(content: bytes) -> ParsedMatrix:
    if not content.startswith(b"PK"):
        raise ValueError("Excel workbook (.xlsx) is required; this file is not a valid .xlsx")
    from zipfile import BadZipFile

    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError, KeyError) as exc:
        raise ValueError("Could not read the Excel training matrix") from exc
    try:
        sheet = workbook.active
        rows = [[_cell_str(value) for value in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()
    return parse_training_matrix_rows(rows)


def parse_training_matrix_upload(filename: str | None, content: bytes) -> ParsedMatrix:
    name = (filename or "training-matrix.csv").lower()
    if name.endswith(".xlsx"):
        return parse_training_matrix_xlsx(content)
    if name.endswith(".csv"):
        return parse_training_matrix_csv(content)
    raise ValueError(
        "Upload the Atlas or Citation Training Matrix Report as CSV or Excel (.csv / .xlsx). " "PDF is not accepted."
    )

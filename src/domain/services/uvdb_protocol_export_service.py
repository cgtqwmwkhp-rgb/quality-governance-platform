"""UVDB Verify B2 protocol pack export (JSON + XLSX)."""

from __future__ import annotations

import json
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Literal

from openpyxl import Workbook

from src.domain.exceptions import BadRequestError

PACK_VERSION = "uvdb-protocol-1.0"
SUPPORTED_FORMATS = frozenset({"json", "xlsx"})


def build_protocol_structure_payload(sections: list[dict[str, Any]]) -> dict[str, Any]:
    """Shared protocol structure used by GET /protocol and export pack builders."""
    return {
        "protocol_name": "UVDB Verify B2 Audit Protocol",
        "version": "V11.2",
        "reference": "UVDB-QS-003",
        "description": "Comprehensive supply chain qualification audit for UK utilities sector",
        "sections": sections,
        "total_sections": len(sections),
        "scoring": {
            "0": "Non-Compliant - No evidence or systems in place",
            "1": "Partially Compliant - Some evidence but gaps identified",
            "2": "Largely Compliant - Minor improvements needed",
            "3": "Compliant - Full evidence and effective implementation",
        },
        "iso_cross_mapping": {
            "1.1": "ISO 9001:2015 (Quality Management)",
            "1.2": "ISO 45001:2018 (OH&S Management)",
            "1.3": "ISO 14001:2015 (Environmental Management)",
            "2.3": "ISO 27001:2022 (Information Security)",
        },
    }


def build_protocol_pack(
    sections: list[dict[str, Any]],
    *,
    exported_by: str | None = None,
) -> dict[str, Any]:
    """Build an attributable offline protocol pack from static B2 sections."""
    structure = build_protocol_structure_payload(sections)
    return {
        "pack_version": PACK_VERSION,
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "exported_by": exported_by,
        "follow_on_exports": {
            "filled_audit_pack": "not_available",
            "branded_pdf": "not_available",
        },
        **structure,
    }


def _default_filename(fmt: str) -> str:
    date_stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    extension = "json" if fmt == "json" else "xlsx"
    return f"uvdb-protocol-pack-{date_stamp}.{extension}"


def _serialize_iso_mapping(mapping: Any) -> str:
    if not mapping:
        return ""
    if isinstance(mapping, dict):
        return "; ".join(f"{key}: {value}" for key, value in mapping.items())
    return str(mapping)


def build_xlsx_bytes(pack: dict[str, Any]) -> bytes:
    """Flatten protocol pack into an offline-review workbook."""
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Field", "Value"])
    overview_rows = [
        ("pack_version", pack.get("pack_version")),
        ("exported_at", pack.get("exported_at")),
        ("exported_by", pack.get("exported_by")),
        ("protocol_name", pack.get("protocol_name")),
        ("version", pack.get("version")),
        ("reference", pack.get("reference")),
        ("description", pack.get("description")),
        ("total_sections", pack.get("total_sections")),
    ]
    for label, value in overview_rows:
        overview.append([label, value])

    overview.append([])
    overview.append(["Scoring scale", "Definition"])
    for score, definition in (pack.get("scoring") or {}).items():
        overview.append([score, definition])

    overview.append([])
    overview.append(["ISO cross-mapping", "Standard"])
    for question, standard in (pack.get("iso_cross_mapping") or {}).items():
        overview.append([question, standard])

    overview.append([])
    overview.append(["Follow-on export", "Status"])
    for export_name, status in (pack.get("follow_on_exports") or {}).items():
        overview.append([export_name, status])

    sections_sheet = workbook.create_sheet("Sections")
    sections_sheet.append(["section_number", "title", "max_score", "question_count", "iso_mapping"])
    for section in pack.get("sections") or []:
        sections_sheet.append(
            [
                section.get("number"),
                section.get("title"),
                section.get("max_score"),
                len(section.get("questions") or []),
                _serialize_iso_mapping(section.get("iso_mapping")),
            ]
        )

    questions_sheet = workbook.create_sheet("Questions")
    questions_sheet.append(
        [
            "section_number",
            "section_title",
            "question_number",
            "question_text",
            "sub_questions",
            "iso_mapping",
            "evidence",
            "site_applicable",
        ]
    )
    for section in pack.get("sections") or []:
        for question in section.get("questions") or []:
            sub_questions = question.get("sub_questions") or []
            evidence = question.get("evidence") or []
            questions_sheet.append(
                [
                    section.get("number"),
                    section.get("title"),
                    question.get("number"),
                    question.get("text"),
                    "\n".join(sub_questions),
                    _serialize_iso_mapping(question.get("iso_mapping")),
                    "; ".join(evidence),
                    question.get("site_applicable", False),
                ]
            )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_protocol_export(
    sections: list[dict[str, Any]],
    *,
    export_format: Literal["json", "xlsx"] = "json",
    exported_by: str | None = None,
) -> tuple[bytes, str, str]:
    """Return (body, filename, media_type) for the requested export format."""
    fmt = export_format.lower()
    if fmt not in SUPPORTED_FORMATS:
        raise BadRequestError(f"Unsupported export format '{export_format}'. Use json or xlsx.")

    pack = build_protocol_pack(sections, exported_by=exported_by)
    filename = _default_filename(fmt)

    if fmt == "json":
        body = json.dumps(pack, indent=2, default=str).encode("utf-8")
        return body, filename, "application/json"

    return build_xlsx_bytes(pack), filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

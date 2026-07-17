"""Planet Mark MS XLSX year-totals ingest (PM-W1b).

Parses Planet Mark "MS output" Member Copy workbooks and upserts year-level
carbon totals onto an existing CarbonReportingYear.

Authoritative sheets:
  - Scope (Market Based) — Scope 1 / 2 / 3 current tCO₂e
  - Scope (Location Based) — Scope 2 location-based (optional)
  - Total CF — market-based total + employee count (FTE)
"""

from __future__ import annotations

import io
import re
from dataclasses import asdict, dataclass
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.domain.exceptions import BadRequestError, NotFoundError
from src.domain.models.planet_mark import CarbonReportingYear, EmissionSource, Scope3CategoryData
from src.domain.services.planet_mark_service import PlanetMarkService

MAX_XLSX_BYTES = 10 * 1024 * 1024
SHEET_SCOPE_MARKET = "Scope (Market Based)"
SHEET_SCOPE_LOCATION = "Scope (Location Based)"
SHEET_TOTAL_CF = "Total CF"
REQUIRED_SHEETS = (SHEET_SCOPE_MARKET, SHEET_TOTAL_CF)

MS_XLSX_ACTIVITY_TYPE = "ms_xlsx_aggregate"
MS_XLSX_SOURCE_CATEGORY = "ms_xlsx"
YEAR_LABEL_RE = re.compile(r"(YE\d{4})", re.IGNORECASE)


@dataclass(frozen=True)
class ParsedMsXlsxYearTotals:
    scope_1: float
    scope_2_market: float
    scope_2_location: Optional[float]
    scope_3: float
    total_market: float
    average_fte: Optional[float]
    emissions_per_fte: Optional[float]
    workbook_year_label: Optional[str]
    source_filename: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class MsXlsxIngestResult:
    year_id: int
    year_label: str
    totals: ParsedMsXlsxYearTotals
    sources_upserted: int
    message: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "year_id": self.year_id,
            "year_label": self.year_label,
            "scope_1": self.totals.scope_1,
            "scope_2_market": self.totals.scope_2_market,
            "scope_2_location": self.totals.scope_2_location,
            "scope_3": self.totals.scope_3,
            "total_emissions": self.totals.total_market,
            "average_fte": self.totals.average_fte,
            "emissions_per_fte": self.totals.emissions_per_fte,
            "workbook_year_label": self.totals.workbook_year_label,
            "source_filename": self.totals.source_filename,
            "sources_upserted": self.sources_upserted,
            "message": self.message,
        }


class PlanetMarkXlsxIngestService:
    """Parse Planet Mark MS XLSX workbooks and upsert year carbon totals."""

    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    @classmethod
    def parse_xlsx(cls, content: bytes, *, filename: str = "") -> ParsedMsXlsxYearTotals:
        if not content:
            raise BadRequestError("XLSX file is empty")
        if len(content) > MAX_XLSX_BYTES:
            raise BadRequestError(f"XLSX file exceeds {MAX_XLSX_BYTES // (1024 * 1024)} MiB limit")

        try:
            workbook = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001 — surface honest parse errors
            raise BadRequestError(f"Unable to read XLSX workbook: {exc}") from exc

        try:
            missing = [name for name in REQUIRED_SHEETS if name not in workbook.sheetnames]
            if missing:
                raise BadRequestError(
                    "Workbook is not a Planet Mark MS output file — missing required sheet(s): " + ", ".join(missing)
                )

            market = cls._parse_scope_sheet(workbook[SHEET_SCOPE_MARKET], sheet_label=SHEET_SCOPE_MARKET)
            location_scope_2: Optional[float] = None
            if SHEET_SCOPE_LOCATION in workbook.sheetnames:
                location = cls._parse_scope_sheet(workbook[SHEET_SCOPE_LOCATION], sheet_label=SHEET_SCOPE_LOCATION)
                location_scope_2 = location.get("scope_2")

            total_market, average_fte, per_fte = cls._parse_total_cf(workbook[SHEET_TOTAL_CF])

            scope_1 = market["scope_1"]
            scope_2 = market["scope_2"]
            scope_3 = market["scope_3"]
            if scope_1 is None or scope_2 is None or scope_3 is None:
                raise BadRequestError(
                    "Scope (Market Based) must include Scope 1, Scope 2, and Scope 3 current tCO₂e rows"
                )
            if total_market is None or total_market <= 0:
                raise BadRequestError("Total CF market-based total tCO₂e is missing or not positive")

            sum_scopes = scope_1 + scope_2 + scope_3
            if abs(sum_scopes - total_market) > max(0.5, total_market * 0.02):
                raise BadRequestError(
                    "Workbook totals are inconsistent: "
                    f"Scope 1+2+3 ({sum_scopes:.3f}) does not match Total CF market total "
                    f"({total_market:.3f})"
                )

            year_label = cls._infer_year_label(filename)
            return ParsedMsXlsxYearTotals(
                scope_1=float(scope_1),
                scope_2_market=float(scope_2),
                scope_2_location=float(location_scope_2) if location_scope_2 is not None else None,
                scope_3=float(scope_3),
                total_market=float(total_market),
                average_fte=float(average_fte) if average_fte is not None else None,
                emissions_per_fte=float(per_fte) if per_fte is not None else None,
                workbook_year_label=year_label,
                source_filename=filename or "upload.xlsx",
            )
        finally:
            workbook.close()

    async def ingest(
        self,
        *,
        year_id: int,
        tenant_id: int | None,
        content: bytes,
        filename: str,
    ) -> MsXlsxIngestResult:
        year = (
            await self.db.execute(
                select(CarbonReportingYear).where(
                    CarbonReportingYear.id == year_id,
                    CarbonReportingYear.tenant_id == tenant_id,
                )
            )
        ).scalar_one_or_none()
        if year is None:
            raise NotFoundError("Reporting year not found")

        parsed = self.parse_xlsx(content, filename=filename)
        if parsed.workbook_year_label and parsed.workbook_year_label.upper() != year.year_label.upper():
            raise BadRequestError(
                f"Workbook year label {parsed.workbook_year_label} does not match selected "
                f"reporting year {year.year_label}. Upload the matching MS XLSX Member Copy."
            )

        real_source = (
            await self.db.execute(
                select(EmissionSource.id).where(
                    EmissionSource.reporting_year_id == year.id,
                    EmissionSource.is_imported_aggregate == False,  # noqa: E712
                )
            )
        ).scalar_one_or_none()
        if real_source is not None:
            raise BadRequestError(
                "This reporting year has manual emission sources. "
                "Remove them before MS XLSX ingest so workbook totals stay the single source of truth."
            )

        for old in (
            (
                await self.db.execute(
                    select(EmissionSource).where(
                        EmissionSource.reporting_year_id == year.id,
                        EmissionSource.is_imported_aggregate == True,  # noqa: E712
                    )
                )
            )
            .scalars()
            .all()
        ):
            await self.db.delete(old)
        await self.db.flush()

        sources_upserted = 0
        scope_rows = (
            ("scope_1", parsed.scope_1, "Scope 1 — Direct Emissions (MS XLSX)"),
            ("scope_2", parsed.scope_2_market, "Scope 2 — Indirect Energy Market (MS XLSX)"),
            ("scope_3", parsed.scope_3, "Scope 3 — Value Chain (MS XLSX)"),
        )
        for scope_key, co2e, source_name in scope_rows:
            if co2e <= 0:
                continue
            self.db.add(
                EmissionSource(  # type: ignore[misc]
                    tenant_id=tenant_id,
                    reporting_year_id=year.id,
                    source_name=source_name,
                    source_category=MS_XLSX_SOURCE_CATEGORY,
                    scope=scope_key,
                    activity_type=MS_XLSX_ACTIVITY_TYPE,
                    activity_value=float(co2e),
                    activity_unit="tCO2e",
                    emission_factor=1.0,
                    emission_factor_unit="tCO2e",
                    emission_factor_source="Planet Mark MS XLSX",
                    co2e_tonnes=float(co2e),
                    data_quality_level="calculated",
                    data_quality_score=3,
                    is_imported_aggregate=True,
                    data_notes=(
                        f"Ingested from Planet Mark MS XLSX ({parsed.source_filename}). "
                        "Replace with per-source data for full detail."
                    ),
                )
            )
            sources_upserted += 1

        if parsed.average_fte is not None and parsed.average_fte > 0:
            year.average_fte = float(parsed.average_fte)
        if parsed.scope_2_location is not None:
            year.scope_2_location = float(parsed.scope_2_location)

        if parsed.scope_3 > 0:
            await self._upsert_scope3_aggregate(year=year, scope_3_total=parsed.scope_3)

        await self.db.flush()
        await PlanetMarkService.recalculate_year_totals(self.db, year)

        # Workbook Total CF is the product SSOT — keep it authoritative after recalculate.
        year.total_emissions = float(parsed.total_market)
        year.scope_1_total = float(parsed.scope_1)
        year.scope_2_market = float(parsed.scope_2_market)
        year.scope_3_total = float(parsed.scope_3)
        if year.average_fte and year.average_fte > 0:
            year.emissions_per_fte = year.total_emissions / year.average_fte
        elif parsed.emissions_per_fte is not None:
            year.emissions_per_fte = float(parsed.emissions_per_fte)

        await self.db.flush()

        return MsXlsxIngestResult(
            year_id=year.id,
            year_label=year.year_label,
            totals=parsed,
            sources_upserted=sources_upserted,
            message="MS XLSX year totals ingested",
        )

    async def _upsert_scope3_aggregate(self, *, year: CarbonReportingYear, scope_3_total: float) -> None:
        other_cat = (
            await self.db.execute(
                select(Scope3CategoryData).where(
                    Scope3CategoryData.reporting_year_id == year.id,
                    Scope3CategoryData.category_number == 15,
                )
            )
        ).scalar_one_or_none()
        note = "Aggregate Scope 3 from Planet Mark MS XLSX. Break down by category to improve data quality."
        if other_cat is not None:
            other_cat.total_co2e = float(scope_3_total)
            other_cat.is_measured = True
            other_cat.calculation_method = "spend_based"
            other_cat.category_description = note
            return
        self.db.add(
            Scope3CategoryData(  # type: ignore[misc]
                reporting_year_id=year.id,
                tenant_id=year.tenant_id,
                category_number=15,
                category_name="Other (Unattributed Scope 3)",
                category_description=note,
                is_relevant=True,
                is_measured=True,
                total_co2e=float(scope_3_total),
                calculation_method="spend_based",
            )
        )

    @staticmethod
    def _infer_year_label(filename: str) -> Optional[str]:
        match = YEAR_LABEL_RE.search(filename or "")
        if not match:
            return None
        return match.group(1).upper()

    @staticmethod
    def _as_float(value: object) -> Optional[float]:
        if value is None or value == "":
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(",", "")
        if not text or text == "-":
            return None
        try:
            return float(text)
        except ValueError:
            return None

    @classmethod
    def _parse_scope_sheet(cls, sheet: Any, *, sheet_label: str) -> dict[str, Optional[float]]:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise BadRequestError(f"Sheet '{sheet_label}' is empty")

        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        try:
            scope_idx = header.index("Scope")
        except ValueError as exc:
            raise BadRequestError(f"Sheet '{sheet_label}' missing 'Scope' column") from exc

        current_idx = None
        for candidate in ("tCo2e_current", "tCO2e_current", "ex_tCo2e_current"):
            if candidate in header:
                current_idx = header.index(candidate)
                break
        if current_idx is None:
            raise BadRequestError(f"Sheet '{sheet_label}' missing current tCO₂e column (tCo2e_current)")

        out: dict[str, Optional[float]] = {"scope_1": None, "scope_2": None, "scope_3": None}
        for row in rows[1:]:
            if not row or scope_idx >= len(row):
                continue
            label = str(row[scope_idx] or "").strip().lower()
            if label == "scope 1":
                out["scope_1"] = cls._as_float(row[current_idx] if current_idx < len(row) else None)
            elif label == "scope 2":
                out["scope_2"] = cls._as_float(row[current_idx] if current_idx < len(row) else None)
            elif label == "scope 3":
                out["scope_3"] = cls._as_float(row[current_idx] if current_idx < len(row) else None)
        return out

    @classmethod
    def _parse_total_cf(cls, sheet: Any) -> tuple[Optional[float], Optional[float], Optional[float]]:
        mode: Optional[str] = None
        total_market: Optional[float] = None
        average_fte: Optional[float] = None
        per_fte: Optional[float] = None

        for row in sheet.iter_rows(values_only=True):
            cells = [c for c in row if c is not None]
            joined = " ".join(str(c) for c in cells)
            if "Total Market Based" in joined:
                mode = "market"
                continue
            if "Total Location Based" in joined:
                mode = "location"
                continue
            if not row:
                continue
            label = str(row[0]).strip() if row[0] is not None else ""
            if mode != "market":
                continue
            # Columns: 0 Scope, 4 base, 6 current (ex_tCo2e_current)
            current = cls._as_float(row[6] if len(row) > 6 else None)
            if label == "Total" and current is not None:
                total_market = current
            elif label == "No. employees" and current is not None:
                average_fte = current
            elif label == "Total per employee" and current is not None:
                per_fte = current

        return total_market, average_fte, per_fte

"""Parse Plantexpand H&S Incident Model workbook sheets into typed rows."""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Optional

from openpyxl import load_workbook

from src.domain.services.hs_rta_normalization import normalize_rta_collision_type

SOURCE_FORM_ID = "hs_excel_v2"


def parse_yn(value: Any) -> Optional[bool]:
    if value is None or str(value).strip() == "":
        return None
    token = str(value).strip().lower()
    if token in {"y", "yes", "true", "1"}:
        return True
    if token in {"n", "no", "false", "0"}:
        return False
    return None


def _cell(row: tuple[Any, ...], idx: int) -> Any:
    return row[idx] if idx < len(row) else None


def _as_aware(value: Any) -> Optional[datetime]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    return None


def _text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def route_incident_log_type(raw_type: str) -> str:
    key = raw_type.strip().lower()
    if key.startswith("injury"):
        return "incident"
    if "near miss" in key:
        return "near_miss"
    if "complaint" in key:
        return "complaint"
    if key in {"rta", "rtc", "road traffic"} or key.startswith("rta"):
        return "rta"
    return "unknown"


def external_key(sheet: str, excel_id: Any) -> str:
    return f"excel:{sheet}:{excel_id}"


def parse_hs_workbook(content: bytes) -> dict[str, Any]:
    """Return structured rows from Incident Log + RTA Log sheets."""
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    incident_rows: list[dict[str, Any]] = []
    rta_rows: list[dict[str, Any]] = []
    warnings: list[str] = []

    if "Incident Log" in workbook.sheetnames:
        ws = workbook["Incident Log"]
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            excel_id = _cell(row, 0)
            if excel_id is None and not any(row):
                continue
            event_date = _as_aware(_cell(row, 1))
            raw_type = _text(_cell(row, 5))
            module = route_incident_log_type(raw_type) if raw_type else "unknown"
            if not event_date:
                warnings.append(f"Incident Log row {idx}: skipped (no date)")
                continue
            if module == "unknown":
                warnings.append(f"Incident Log row {idx}: unknown type {raw_type!r}")
                continue
            status_raw = _text(_cell(row, 16)).lower()
            closed = status_raw == "closed"
            incident_rows.append(
                {
                    "sheet": "incident_log",
                    "excel_id": excel_id,
                    "external_key": external_key("incident_log", excel_id),
                    "module": module,
                    "event_date": event_date,
                    "reporter": _text(_cell(row, 3)) or "Unknown",
                    "customer": _text(_cell(row, 4)),
                    "raw_type": raw_type,
                    "person_involved": _text(_cell(row, 6)),
                    "role_location": _text(_cell(row, 7)),
                    "description": _text(_cell(row, 8)) or "(no description)",
                    "is_injury": parse_yn(_cell(row, 9)),
                    "body_part": _text(_cell(row, 10)),
                    "medical_assistance": parse_yn(_cell(row, 11)),
                    "is_riddor": parse_yn(_cell(row, 12)),
                    "is_lti": parse_yn(_cell(row, 13)),
                    "is_minor_injury": parse_yn(_cell(row, 14)),
                    "is_hipo": parse_yn(_cell(row, 15)),
                    "closed": closed,
                    "notes": _text(_cell(row, 17)),
                }
            )

    if "RTA Log" in workbook.sheetnames:
        ws = workbook["RTA Log"]
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            excel_id = _cell(row, 0)
            if excel_id is None and not any(row):
                continue
            event_date = _as_aware(_cell(row, 1))
            if not event_date:
                warnings.append(f"RTA Log row {idx}: skipped (no date)")
                continue
            rta_rows.append(
                {
                    "sheet": "rta_log",
                    "excel_id": excel_id,
                    "external_key": external_key("rta_log", excel_id),
                    "module": "rta",
                    "event_date": event_date,
                    "employee": _text(_cell(row, 3)) or "Unknown",
                    "vehicle_reg": _text(_cell(row, 4)),
                    "time": _text(_cell(row, 5)),
                    "location": _text(_cell(row, 6)) or "Unknown",
                    "collision_type": normalize_rta_collision_type(_text(_cell(row, 7))),
                    "damage": _text(_cell(row, 8)),
                    "drivable": parse_yn(_cell(row, 9)),
                    "weather": _text(_cell(row, 10)),
                    "road_conditions": _text(_cell(row, 11)),
                    "emergency_services": parse_yn(_cell(row, 12)),
                    "third_party_injury": parse_yn(_cell(row, 13)),
                    "employee_injured": parse_yn(_cell(row, 14)),
                    "is_lti": parse_yn(_cell(row, 15)),
                    "is_riddor": parse_yn(_cell(row, 16)),
                    "notes": _text(_cell(row, 17)),
                }
            )

    workbook.close()
    return {"incident_log": incident_rows, "rta_log": rta_rows, "warnings": warnings}

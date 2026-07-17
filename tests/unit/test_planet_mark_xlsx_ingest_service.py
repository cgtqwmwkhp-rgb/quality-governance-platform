"""Unit tests for Planet Mark MS XLSX year-totals ingest (PM-W1b)."""

from __future__ import annotations

import io
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import Workbook

from src.domain.exceptions import BadRequestError
from src.domain.services.planet_mark_xlsx_ingest_service import PlanetMarkXlsxIngestService


def _build_ms_xlsx(
    *,
    scope_1: float = 477.439,
    scope_2: float = 2.049,
    scope_3: float = 174.971,
    total_market: float = 654.459,
    fte: float = 94.375,
    scope_2_location: float = 10.876,
) -> bytes:
    wb = Workbook()
    # openpyxl creates a default sheet — rename/rebuild required sheets
    default = wb.active
    assert default is not None
    wb.remove(default)

    market = wb.create_sheet("Scope (Market Based)")
    market.append(
        [
            "Scope",
            "tCo2e_base",
            "ex_tCo2e_base",
            "Consumption Amount_base",
            "normalised_extCo2e_base",
            "tCo2e_current",
            "ex_tCo2e_current",
        ]
    )
    market.append(["Scope 1", 0, 0, 0, 0, scope_1, scope_1])
    market.append(["Scope 1 & 2", 0, 0, 0, 0, scope_1 + scope_2, scope_1 + scope_2])
    market.append(["Scope 2", 0, 0, 0, 0, scope_2, scope_2])
    market.append(["Scope 3", 0, 0, 0, 0, scope_3, scope_3])

    location = wb.create_sheet("Scope (Location Based)")
    location.append(
        [
            "Scope",
            "tCo2e_base",
            "ex_tCo2e_base",
            "Consumption Amount_base",
            "normalised_extCo2e_base",
            "tCo2e_current",
            "ex_tCo2e_current",
        ]
    )
    location.append(["Scope 1", 0, 0, 0, 0, scope_1, scope_1])
    location.append(["Scope 2", 0, 0, 0, 0, scope_2_location, scope_2_location])
    location.append(["Scope 3", 0, 0, 0, 0, scope_3, scope_3])

    total = wb.create_sheet("Total CF")
    total.append(
        [
            "Scope",
            "Emission Source",
            "Unit",
            "Consumption Amount_base",
            "ex_tCo2e_base",
            "Consumption Amount_current",
            "ex_tCo2e_current",
        ]
    )
    total.append([None, None, None, "Total Market Based", None, None, None])
    total.append(["Total", None, "tCO₂e", None, 100, None, total_market])
    total.append(["No. employees", None, "Number", None, 50, None, fte])
    total.append(["Total per employee", None, "tCO₂e", None, 2, None, total_market / fte])
    total.append([None, None, None, "Total Location Based", None, None, None])
    total.append(["Total", None, "tCO₂e", None, 110, None, total_market + 9])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_xlsx_reads_market_scope_and_total_cf():
    content = _build_ms_xlsx()
    parsed = PlanetMarkXlsxIngestService.parse_xlsx(
        content, filename="Plantexpand_Planet Mark_MS output - YE2024 - Member Copy.xlsx"
    )
    assert parsed.workbook_year_label == "YE2024"
    assert parsed.scope_1 == pytest.approx(477.439, rel=1e-3)
    assert parsed.scope_2_market == pytest.approx(2.049, rel=1e-3)
    assert parsed.scope_2_location == pytest.approx(10.876, rel=1e-3)
    assert parsed.scope_3 == pytest.approx(174.971, rel=1e-3)
    assert parsed.total_market == pytest.approx(654.459, rel=1e-3)
    assert parsed.average_fte == pytest.approx(94.375, rel=1e-3)


def test_parse_xlsx_rejects_missing_required_sheet():
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Input"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(BadRequestError, match="missing required sheet"):
        PlanetMarkXlsxIngestService.parse_xlsx(buf.getvalue(), filename="bad.xlsx")


def test_parse_xlsx_rejects_inconsistent_totals():
    content = _build_ms_xlsx(total_market=100.0)
    with pytest.raises(BadRequestError, match="inconsistent"):
        PlanetMarkXlsxIngestService.parse_xlsx(content, filename="YE2024.xlsx")


def test_parse_real_ye2024_member_copy_when_present():
    from pathlib import Path

    path = Path("/Users/davidharris/Downloads/Plantexpand_Planet Mark_MS output - YE2024 - Member Copy.xlsx")
    if not path.exists():
        pytest.skip("YE2024 MS XLSX not available on this machine")
    parsed = PlanetMarkXlsxIngestService.parse_xlsx(path.read_bytes(), filename=path.name)
    assert parsed.workbook_year_label == "YE2024"
    assert parsed.total_market == pytest.approx(654.458, rel=1e-3)
    assert parsed.scope_1 == pytest.approx(477.439, rel=1e-3)
    assert parsed.average_fte == pytest.approx(94.375, rel=1e-3)


@pytest.mark.asyncio
async def test_ingest_rejects_year_label_mismatch():
    db = AsyncMock()
    year = SimpleNamespace(id=7, year_label="YE2025", tenant_id=1, average_fte=0.0)
    year_result = MagicMock()
    year_result.scalar_one_or_none.return_value = year
    real_result = MagicMock()
    real_result.scalar_one_or_none.return_value = None
    db.execute = AsyncMock(side_effect=[year_result, real_result])

    service = PlanetMarkXlsxIngestService(db)
    content = _build_ms_xlsx()
    with pytest.raises(BadRequestError, match="does not match"):
        await service.ingest(
            year_id=7,
            tenant_id=1,
            content=content,
            filename="Plantexpand_Planet Mark_MS output - YE2024 - Member Copy.xlsx",
        )


@pytest.mark.asyncio
async def test_ingest_upserts_aggregates_and_sets_totals():
    db = AsyncMock()
    year = SimpleNamespace(
        id=7,
        year_label="YE2024",
        tenant_id=1,
        average_fte=0.0,
        scope_2_location=0.0,
        scope_1_total=0.0,
        scope_2_market=0.0,
        scope_3_total=0.0,
        total_emissions=0.0,
        emissions_per_fte=0.0,
    )
    year_result = MagicMock()
    year_result.scalar_one_or_none.return_value = year
    real_result = MagicMock()
    real_result.scalar_one_or_none.return_value = None
    old_aggs = MagicMock()
    old_aggs.scalars.return_value.all.return_value = []
    scope3_result = MagicMock()
    scope3_result.scalar_one_or_none.return_value = None
    db.execute = AsyncMock(side_effect=[year_result, real_result, old_aggs, scope3_result])
    db.flush = AsyncMock()
    db.add = MagicMock()
    db.delete = AsyncMock()

    service = PlanetMarkXlsxIngestService(db)
    content = _build_ms_xlsx()

    with patch(
        "src.domain.services.planet_mark_xlsx_ingest_service.PlanetMarkService.recalculate_year_totals",
        new=AsyncMock(),
    ):
        result = await service.ingest(
            year_id=7,
            tenant_id=1,
            content=content,
            filename="Plantexpand_Planet Mark_MS output - YE2024 - Member Copy.xlsx",
        )

    assert result.year_id == 7
    assert result.sources_upserted == 3
    assert year.total_emissions == pytest.approx(654.459, rel=1e-3)
    assert year.average_fte == pytest.approx(94.375, rel=1e-3)
    assert db.add.call_count >= 3

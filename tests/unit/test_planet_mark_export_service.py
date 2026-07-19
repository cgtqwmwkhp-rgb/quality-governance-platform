"""Unit tests for Planet Mark export pack assembly."""

from __future__ import annotations

from datetime import datetime
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import load_workbook

from src.domain.services.planet_mark_export_service import (
    PlanetMarkExportService,
    build_hotspot_initiatives,
    export_pack_filename,
)


class TestBuildHotspotInitiatives:
    def test_ranks_measured_categories_and_applies_reduction_tiers(self):
        categories = [
            {"number": 1, "name": "Purchased goods", "is_measured": True, "total_co2e": 10, "percentage": 15},
            {"number": 6, "name": "Travel", "is_measured": True, "total_co2e": 40, "percentage": 80},
            {"number": 2, "name": "Capital", "is_measured": False, "total_co2e": 99, "percentage": 0},
        ]
        initiatives = build_hotspot_initiatives(categories, total_co2e=50, limit=5)
        assert len(initiatives) == 2
        assert initiatives[0]["categoryNumber"] == 6
        assert initiatives[0]["suggestedReductionPercent"] == 10
        assert initiatives[1]["categoryNumber"] == 1
        assert initiatives[1]["suggestedReductionPercent"] == 5


class TestExportPackFilename:
    def test_sanitizes_year_label(self):
        name = export_pack_filename("YE 2026/A", "json")
        assert name.startswith("planet-mark-export-YE_2026_A-")
        assert name.endswith(".json")


@pytest.mark.asyncio
async def test_assemble_pack_shape_and_xlsx_bytes():
    year = SimpleNamespace(
        id=7,
        year_label="YE2026",
        year_number=2026,
        period_start=datetime(2026, 1, 1),
        period_end=datetime(2026, 12, 31),
        average_fte=12.0,
        total_emissions=100.0,
        emissions_per_fte=8.3,
        scope_1_total=40.0,
        scope_2_market=30.0,
        scope_3_total=30.0,
        overall_data_quality=12,
        certification_status="in_progress",
        is_baseline_year=False,
    )
    scope3_row = SimpleNamespace(
        category_number=6,
        category_name="Business travel",
        is_measured=True,
        total_co2e=30.0,
    )
    action = SimpleNamespace(
        id=1,
        action_id="ACT-1",
        action_title="Reduce travel",
        achievable_owner="Ops",
        time_bound=datetime(2026, 12, 31),
        status="in_progress",
        progress_percent=40,
        expected_reduction_pct=5.0,
    )

    db = MagicMock()
    year_result = MagicMock()
    year_result.scalar_one_or_none.return_value = year
    scope3_result = MagicMock()
    scope3_result.scalars.return_value.all.return_value = [scope3_row]
    actions_result = MagicMock()
    actions_result.scalars.return_value.all.return_value = [action]
    db.execute = AsyncMock(side_effect=[year_result, scope3_result, actions_result])

    service = PlanetMarkExportService(db)
    pack = await service.assemble_pack(year_id=7, tenant_id=1)

    assert pack["export_kind"] == "json_pack"
    assert pack["reporting_year"]["year_label"] == "YE2026"
    assert pack["scope3_categories"][0]["number"] == 6
    assert pack["improvement_actions"][0]["action_id"] == "ACT-1"
    assert pack["hotspot_initiatives"][0]["categoryNumber"] == 6

    json_bytes = service.build_json_bytes(pack)
    assert b'"year_label": "YE2026"' in json_bytes

    xlsx_bytes = service.build_xlsx_bytes(pack)
    workbook = load_workbook(filename=__import__("io").BytesIO(xlsx_bytes))
    assert workbook.sheetnames == ["Summary", "Scope 3", "Actions", "Hotspot initiatives"]
    assert workbook["Summary"]["A1"].value == "Planet Mark export pack"


@pytest.mark.asyncio
async def test_assemble_pack_missing_year_raises():
    db = MagicMock()
    year_result = MagicMock()
    year_result.scalar_one_or_none.return_value = None
    db.execute = AsyncMock(return_value=year_result)

    service = PlanetMarkExportService(db)
    with pytest.raises(ValueError, match="Reporting year not found"):
        await service.assemble_pack(year_id=99, tenant_id=1)

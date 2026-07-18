"""Unit tests for RR-W4 Enterprise Risk Register XLSX import."""

from __future__ import annotations

import io
from datetime import datetime
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import Workbook

from src.domain.exceptions import BadRequestError, ValidationError
from src.domain.services.risk_register_import_service import RiskRegisterImportService


def _workbook_bytes(
    rows: list[list[object]],
    *,
    action_plan_rows: list[list[object]] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Register"
    for row in rows:
        ws.append(row)
    if action_plan_rows is not None:
        ap = wb.create_sheet("Action Plan")
        for row in action_plan_rows:
            ap.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


ACTION_PLAN_HEADER = [
    "Action ID",
    "Linked Risk Ref",
    "Action Description",
    "Owner",
    "Cost (GBP)",
    "Deadline",
    "Status",
    "Progress Notes",
    "MatchKey",
]


def _valid_action_row(action_id: str = "A001", risk_ref: str = "PELR1") -> list[object]:
    return [
        action_id,
        risk_ref,
        "Create a process guide to support the process handover",
        "James Anang",
        0,
        datetime(2025, 9, 29),
        "In progress",
        None,
        f"{risk_ref}-1",
    ]


HEADER = [
    "Ref",
    "Date Identified",
    "Risk Title",
    "Risk Description",
    "Root Causes",
    "Category",
    "Risk Owner",
    "Gross Impact (1-5)",
    "Gross Likelihood (1-5)",
    "Gross Score",
    "Gross RAG",
    "Existing Controls",
    "Control Effectiveness",
    "Net Impact (1-5)",
    "Net Likelihood (1-5)",
    "Net Score",
    "Net RAG",
    "Trend",
    "Status",
    "Last Reviewed",
    "Next Review",
    "Comments",
]


def _valid_data_row(ref: str = "PELR1") -> list[object]:
    return [
        ref,
        datetime(2024, 4, 8),
        "Thames Water TWOSA Application",
        "It was noted that PE had not fully met the requirements for TWOSA applications with Thames Water.",
        "Lack of clear ownership of the application process.",
        "Compliance",
        "James Anang (H&S)",
        3,
        4,
        12,
        "Medium",
        "Live TWOSA tracker implemented.",
        "Effective",
        2,
        2,
        4,
        "Low",
        "Decreasing",
        "Open",
        datetime(2026, 6, 1),
        datetime(2026, 9, 1),
        "Reviewed in June.",
    ]


@pytest.fixture
def valid_xlsx() -> bytes:
    return _workbook_bytes([HEADER, _valid_data_row()])


@pytest.fixture
def invalid_ref_xlsx() -> bytes:
    return _workbook_bytes([HEADER, _valid_data_row("BAD-1")])


def _service_with_existing(existing: dict | None = None) -> RiskRegisterImportService:
    db = MagicMock()
    service = RiskRegisterImportService(db)
    existing_map = existing or {}

    async def _existing_references(references: set[str], tenant_id: int):  # noqa: ARG001
        return existing_map

    setattr(service, "_existing_references", _existing_references)
    setattr(
        service,
        "_appetite_threshold",
        AsyncMock(return_value=12),
    )
    return service


def test_parse_xlsx_reads_risk_register_sheet(valid_xlsx: bytes):
    rows = RiskRegisterImportService.parse_xlsx(valid_xlsx)
    assert len(rows) == 1
    assert rows[0]["reference"] == "PELR1"
    assert rows[0]["title"] == "Thames Water TWOSA Application"


def test_parse_xlsx_requires_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.append(["Ref"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(BadRequestError, match="Risk Register"):
        RiskRegisterImportService.parse_xlsx(buf.getvalue())


@pytest.mark.asyncio
async def test_dry_run_reports_create(valid_xlsx: bytes):
    service = _service_with_existing()
    report = await service.dry_run(valid_xlsx, tenant_id=1)
    assert report.ok is True
    assert report.total_rows == 1
    assert report.creates == 1
    assert report.updates == 0
    assert report.preview[0]["reference"] == "PELR1"
    assert report.action_plan_skipped is True


@pytest.mark.asyncio
async def test_dry_run_maps_action_plan_to_capa():
    xlsx = _workbook_bytes(
        [HEADER, _valid_data_row()],
        action_plan_rows=[ACTION_PLAN_HEADER, _valid_action_row()],
    )
    service = _service_with_existing()
    setattr(service, "_existing_capa_by_source_refs", AsyncMock(return_value={}))
    report = await service.dry_run(xlsx, tenant_id=1)
    assert report.ok is True
    assert report.action_plan_skipped is False
    assert report.action_plan_total_rows == 1
    assert report.action_plan_creates == 1
    assert report.action_plan_preview[0]["action_id"] == "A001"
    assert report.action_plan_preview[0]["risk_reference"] == "PELR1"


@pytest.mark.asyncio
async def test_dry_run_reports_update(valid_xlsx: bytes):
    existing_risk = SimpleNamespace(id=99, reference="PELR1")
    service = _service_with_existing({"PELR1": existing_risk})
    report = await service.dry_run(valid_xlsx, tenant_id=1)
    assert report.ok is True
    assert report.creates == 0
    assert report.updates == 1
    assert report.preview[0]["action"] == "update"


@pytest.mark.asyncio
async def test_dry_run_rejects_invalid_reference(invalid_ref_xlsx: bytes):
    service = _service_with_existing()
    report = await service.dry_run(invalid_ref_xlsx, tenant_id=1)
    assert report.ok is False
    assert report.error_rows == 1
    assert report.errors[0].code == "INVALID_REFERENCE"


@pytest.mark.asyncio
async def test_commit_blocks_on_validation_errors(invalid_ref_xlsx: bytes):
    service = _service_with_existing()
    with pytest.raises(ValidationError, match="validation failed"):
        await service.commit(invalid_ref_xlsx, user_id=7, tenant_id=1)


@pytest.mark.asyncio
async def test_commit_creates_risk(valid_xlsx: bytes):
    db = MagicMock()
    db.flush = AsyncMock()
    db.commit = AsyncMock()
    service = RiskRegisterImportService(db)
    setattr(service, "_existing_references", AsyncMock(return_value={}))
    setattr(service, "_appetite_threshold", AsyncMock(return_value=12))

    result = await service.commit(valid_xlsx, user_id=7, tenant_id=1)
    assert result.created_count == 1
    assert result.updated_count == 0
    db.add.assert_called_once()
    db.commit.assert_awaited_once()

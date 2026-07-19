"""Unit tests for UVDB protocol pack export builder."""

from __future__ import annotations

import json
from io import BytesIO

import pytest
from openpyxl import load_workbook

from src.domain.exceptions import BadRequestError
from src.domain.services.uvdb_protocol_export_service import (
    PACK_VERSION,
    build_protocol_export,
    build_protocol_pack,
    build_protocol_structure_payload,
)

SAMPLE_SECTIONS = [
    {
        "number": "1",
        "title": "System Assurance and Compliance",
        "max_score": 21,
        "iso_mapping": {"9001": "4-5"},
        "questions": [
            {
                "number": "1.1",
                "text": "Can the company demonstrate quality systems?",
                "sub_questions": ["Is certification held?"],
                "iso_mapping": {"9001": ["4.4"]},
                "evidence": ["ISO 9001 certificate"],
            }
        ],
    }
]


def test_build_protocol_structure_payload_matches_protocol_contract() -> None:
    payload = build_protocol_structure_payload(SAMPLE_SECTIONS)

    assert payload["protocol_name"] == "UVDB Verify B2 Audit Protocol"
    assert payload["version"] == "V11.2"
    assert payload["reference"] == "UVDB-QS-003"
    assert payload["total_sections"] == 1
    assert payload["sections"] == SAMPLE_SECTIONS
    assert "3" in payload["scoring"]
    assert payload["iso_cross_mapping"]["1.1"].startswith("ISO 9001")


def test_build_protocol_pack_includes_attribution_and_follow_on_honesty() -> None:
    pack = build_protocol_pack(SAMPLE_SECTIONS, exported_by="auditor@example.com")

    assert pack["pack_version"] == PACK_VERSION
    assert pack["exported_by"] == "auditor@example.com"
    assert pack["exported_at"]
    assert pack["follow_on_exports"]["filled_audit_pack"] == "not_available"
    assert pack["follow_on_exports"]["branded_pdf"] == "not_available"
    assert pack["sections"] == SAMPLE_SECTIONS


def test_build_protocol_export_json_attachment_shape() -> None:
    body, filename, media_type = build_protocol_export(
        SAMPLE_SECTIONS,
        export_format="json",
        exported_by="auditor@example.com",
    )

    assert media_type == "application/json"
    assert filename.endswith(".json")

    pack = json.loads(body.decode("utf-8"))
    assert pack["pack_version"] == PACK_VERSION
    assert pack["exported_by"] == "auditor@example.com"
    assert pack["sections"][0]["questions"][0]["number"] == "1.1"


def test_build_protocol_export_xlsx_contains_sections_and_questions() -> None:
    body, filename, media_type = build_protocol_export(
        SAMPLE_SECTIONS,
        export_format="xlsx",
        exported_by="auditor@example.com",
    )

    assert media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert filename.endswith(".xlsx")

    workbook = load_workbook(BytesIO(body))
    assert workbook.sheetnames == ["Overview", "Sections", "Questions"]

    sections_rows = list(workbook["Sections"].iter_rows(values_only=True))
    assert sections_rows[0] == (
        "section_number",
        "title",
        "max_score",
        "question_count",
        "iso_mapping",
    )
    assert sections_rows[1][0] == "1"
    assert sections_rows[1][3] == 1

    question_rows = list(workbook["Questions"].iter_rows(values_only=True))
    assert question_rows[1][2] == "1.1"
    assert question_rows[1][3] == "Can the company demonstrate quality systems?"


def test_build_protocol_export_rejects_unknown_format() -> None:
    with pytest.raises(BadRequestError, match="Unsupported export format"):
        build_protocol_export(SAMPLE_SECTIONS, export_format="pdf")  # type: ignore[arg-type]

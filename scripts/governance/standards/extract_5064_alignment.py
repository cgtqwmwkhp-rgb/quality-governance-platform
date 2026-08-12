#!/usr/bin/env python3
"""Turn the PEL-HSEQ-5064 workbook into the checked-in alignment payload.

The workbook is not in the repository — it is a Level 5 report that lives in the
document estate. This script reads it and writes
``specs/standards/pel-hseq-5064-alignment-v1.0.json``, which *is* checked in, so
tests and the seed path never need the spreadsheet.

Usage::

    python -m scripts.governance.standards.extract_5064_alignment \\
        --workbook "/path/to/PEL-HSEQ-5064 Standards Alignment Matrix v1.0.xlsx"

What it does not do
-------------------
It does not infer alignments. Every pair verdict written comes either from the
row's own Verdict column or from :data:`PAIR_OVERRIDES`, and each override quotes
the sentence in the workbook that names the subset relationship explicitly.
Inventing a NEAR where the source says DIFFERENT is the exact failure this matrix
exists to prevent, so an alignment the source only implies is left as the row
verdict says.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Optional

SHEET_CLAUSES = "02 Clause alignment 4 to 10"
SHEET_EXACT = "06 EXACT alignments"
SHEET_TRAPS = "07 The traps"

#: Workbook column order on sheet 02, mapped to the framework ids the matrix
#: chrome and the cell-aggregate service already use.
CLAUSE_SHEET_FRAMEWORKS: tuple[tuple[int, str], ...] = (
    (2, "9001"),
    (3, "14001"),
    (4, "45001"),
    (5, "27001"),
    (6, "22301"),
)

#: Constructionline is referenced by the source but deliberately excluded from
#: the programme catalogue (Wave 1 PR-A made the same call in
#: ``standardsMatrixFilters.ts``). Recorded on the version row so the omission
#: reads as a decision rather than as missing data.
EXCLUDED_FRAMEWORKS: tuple[str, ...] = ("constructionline",)

#: Cells that mean "this framework does not have this clause".
_ABSENT_TOKENS = frozenset({"", "n/a", "na", "none", "-", "—"})

#: Cells whose text says the requirement lives somewhere else entirely
#: ("within 8.1.3 management of change"). The clause is present but relocated,
#: which is itself part of why the row is a trap, so the text is kept as the
#: clause label and the clause key falls back to the row reference.
_RELOCATED_PREFIX = "within"

#: Words that make the leading number the *head of a span* rather than the clause
#: itself: "7.5.1 to 7.5.3", "9.2.1 and 9.2.2", "8.1 including 8.1.2 and 8.1.3".
#: A span of sub-clauses belongs to the row's own clause, so the key must stay on
#: the row axis ("7.5"), not drift to the first sub-clause ("7.5.1") where no
#: matrix lookup would ever find it.
_SPAN_CONNECTIVES = ("to ", "and ", "including ", "&")

_VERDICTS = ("EXACT", "NEAR", "DIFFERENT", "UNIQUE")

#: Pair-level verdicts that override the row verdict, each keyed by row and
#: justified by the workbook sentence that states the subset relationship. The
#: note is stored on the edge, so an auditor sees the source's own words.
PAIR_OVERRIDES: dict[str, tuple[dict[str, Any], ...]] = {
    "6.1.3": (
        {
            "pair": ("14001", "45001"),
            "verdict": "NEAR",
            "note": (
                "ISO 14001 6.1.3 and ISO 45001 6.1.3 are near identical and ONE "
                "register satisfies both."
            ),
        },
        # The row verdict is NEAR, but the source says 27001 reuses the number for
        # something unrelated and 22301 does not have it as a legal clause at all.
        {
            "pair": ("14001", "27001"),
            "verdict": "DIFFERENT",
            "note": (
                "ISO 27001 uses the same clause number for something completely "
                "unrelated, risk treatment and the Statement of Applicability."
            ),
        },
        {
            "pair": ("45001", "27001"),
            "verdict": "DIFFERENT",
            "note": (
                "ISO 27001 uses the same clause number for something completely "
                "unrelated, risk treatment and the Statement of Applicability."
            ),
        },
        {
            "pair": ("14001", "22301"),
            "verdict": "DIFFERENT",
            "note": "ISO 22301 6.1.3 is not present as a legal requirements clause.",
        },
        {
            "pair": ("45001", "22301"),
            "verdict": "DIFFERENT",
            "note": "ISO 22301 6.1.3 is not present as a legal requirements clause.",
        },
        {
            "pair": ("27001", "22301"),
            "verdict": "DIFFERENT",
            "note": (
                "ISO 27001 6.1.3 is risk treatment and the Statement of "
                "Applicability; ISO 22301 6.1.3 is not present as a legal clause. "
                "Sharing the number is all they share."
            ),
        },
    ),
    "6.3": (
        {
            "pair": ("9001", "27001"),
            "verdict": "NEAR",
            "note": (
                "ISO 9001 and ISO 27001:2022 have a standalone clause 6.3 requiring "
                "changes to be carried out in a planned manner."
            ),
        },
    ),
    "8.2": (
        {
            "pair": ("14001", "45001"),
            "verdict": "NEAR",
            "note": (
                "ISO 14001 8.2 and ISO 45001 8.2 are NEAR IDENTICAL to each other and "
                "one set of emergency arrangements with a drill record satisfies both."
            ),
        },
    ),
    "9.1.2": (
        {
            "pair": ("14001", "45001"),
            "verdict": "NEAR",
            "note": (
                "ISO 14001 9.1.2 and ISO 45001 9.1.2 are NEAR IDENTICAL: establish, "
                "implement and maintain processes for evaluating fulfilment of legal "
                "and other requirements, and RETAIN DOCUMENTED INFORMATION of the "
                "results."
            ),
        },
    ),
}

#: Cross-framework alignments the workbook states on sheet 06 but which are not
#: rows of the clause-by-clause sheet: the two Investors in People indicators and
#: the two Annex A controls. Written as explicit pairs because that is how the
#: source states them ("Investors in People and ISO 45001 5.4").
SUPPLEMENTARY_ROWS: tuple[dict[str, Any], ...] = (
    {
        "row_key": "iip-3",
        "clause_ref": "IIP 3",
        "title": "Empowering and involving people",
        "verdict": "EXACT",
        "rationale": (
            "The IIP indicator and the 45001 clause ask for the same thing in "
            "different words: people are involved in decisions that affect them. One "
            "consultation mechanism and one record satisfies both."
        ),
        "deliverables": "PEL-PPL-3006, PEL-HSEQ-5026, PEL-PPL-5032",
        "source_sheet": SHEET_EXACT,
        "pairs": (
            {"a": ("iip", "IIP 3"), "b": ("45001", "5.4")},
        ),
    },
    {
        "row_key": "iip-7",
        "clause_ref": "IIP 7",
        "title": "Building capability",
        "verdict": "EXACT",
        "rationale": (
            "The same competence system. IIP asks whether capability is planned and "
            "developed; clause 7.2 asks whether it is determined, achieved and "
            "evidenced. One matrix."
        ),
        "deliverables": "PEL-HSEQ-5029, PEL-PPL-3005, PEL-PPL-4006",
        "source_sheet": SHEET_EXACT,
        "pairs": tuple(
            {"a": ("iip", "IIP 7"), "b": (fw, "7.2")}
            for fw in ("9001", "14001", "45001", "27001", "22301")
        ),
    },
    {
        "row_key": "annexa-5.31",
        "clause_ref": "A.5.31",
        "title": "Legal, statutory, regulatory and contractual requirements",
        "verdict": "EXACT",
        "rationale": (
            "One legal register serves the Annex A control and both management system "
            "clauses. It needs an information security row set, which is why "
            "PEL-IT-5005 should be merged into it rather than kept alongside."
        ),
        "deliverables": "PEL-HSEQ-5056",
        "source_sheet": SHEET_EXACT,
        "pairs": (
            {"a": ("27001", "A.5.31"), "b": ("14001", "6.1.3")},
            {"a": ("27001", "A.5.31"), "b": ("45001", "6.1.3")},
        ),
    },
    {
        "row_key": "annexa-5.33",
        "clause_ref": "A.5.33",
        "title": "Protection of records",
        "verdict": "EXACT",
        "rationale": (
            "The retention and disposition requirement of clause 7.5.3 IS this "
            "control. The register's retention column is the evidence for both."
        ),
        "deliverables": "PEL-HSEQ-3012, PEL-HSEQ-5014, PEL-HSEQ-3027",
        "source_sheet": SHEET_EXACT,
        "pairs": tuple(
            {"a": ("27001", "A.5.33"), "b": (fw, "7.5")}
            for fw in ("9001", "14001", "45001", "27001", "22301")
        ),
    },
)


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _parse_framework_cell(cell: Any, row_clause_ref: str) -> Optional[tuple[str, Optional[str]]]:
    """Return ``(clause_number, label)`` for one framework cell, or None if absent.

    The workbook writes a bare number ("6.1.2"), a number with the framework's own
    subject ("6.1.2 environmental aspects"), a relocation ("within 8.1.3 management
    of change"), a range ("7.5.1 to 7.5.3"), or ``n/a``.
    """
    text = _clean(cell)
    if text.lower() in _ABSENT_TOKENS:
        return None

    if text.lower().startswith(_RELOCATED_PREFIX):
        # "within 8.1.3 management of change" — the requirement exists elsewhere.
        moved = re.search(r"(\d+(?:\.\d+)*)", text)
        return (moved.group(1) if moved else row_clause_ref), text

    match = re.match(r"^((?:A\.)?\d+(?:\.\d+)*)\s*(.*)$", text)
    if not match:
        return row_clause_ref, text
    clause_number = match.group(1)
    remainder = match.group(2).strip()

    is_span = remainder.startswith(_SPAN_CONNECTIVES) or remainder.startswith(".")
    if is_span:
        # "7.5.1 to 7.5.3" on row 7.5 is that row's clause, expressed as a span of
        # its sub-clauses. Only collapse onto the row axis when the span really is
        # inside the row's clause: ISO 27001 answers row 8.1.4 with Annex A controls
        # "A.5.19 to A.5.23", which is a different clause and must stay one.
        if clause_number.startswith(row_clause_ref):
            clause_number = row_clause_ref
        return clause_number, text

    return clause_number, (remainder or None)


def _row_verdict(cell: Any) -> Optional[str]:
    text = _clean(cell).upper()
    for verdict in _VERDICTS:
        if text.startswith(verdict):
            return verdict
    return None


def _extract_clause_rows(worksheet: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
        clause_ref = _clean(values[0] if len(values) > 0 else None)
        verdict = _row_verdict(values[7] if len(values) > 7 else None)
        if not clause_ref or verdict is None:
            continue
        # Sheet 02 has a second block (ISO 22301 clause 8) whose columns mean
        # something else; it carries no Verdict cell, so it is skipped above.
        title = _clean(values[1] if len(values) > 1 else None)

        frameworks: dict[str, dict[str, Any]] = {}
        for column, framework_id in CLAUSE_SHEET_FRAMEWORKS:
            parsed = _parse_framework_cell(
                values[column] if len(values) > column else None,
                clause_ref,
            )
            if parsed is None:
                continue
            clause_number, label = parsed
            frameworks[framework_id] = {"clause_number": clause_number, "label": label}

        rationale = _clean(values[8] if len(values) > 8 else None)
        deliverables = _clean(values[9] if len(values) > 9 else None)

        rows.append(
            {
                "row_key": f"annexsl-{clause_ref}",
                "clause_ref": clause_ref,
                "title": title,
                "verdict": verdict,
                "rationale": rationale or None,
                "deliverables": deliverables or None,
                "frameworks": frameworks,
                "pair_overrides": [
                    {
                        "a": override["pair"][0],
                        "b": override["pair"][1],
                        "verdict": override["verdict"],
                        "note": override["note"],
                    }
                    for override in PAIR_OVERRIDES.get(clause_ref, ())
                ],
                "source_sheet": worksheet.title,
                "source_row": index,
            }
        )
    return rows


def _extract_trap_notes(worksheet: Any) -> dict[str, str]:
    """Sheet 07 restates the DIFFERENT/UNIQUE rows; keep its wording per clause."""
    notes: dict[str, str] = {}
    for values in worksheet.iter_rows(values_only=True):
        clause_ref = _clean(values[0] if len(values) > 0 else None)
        verdict = _row_verdict(values[2] if len(values) > 2 else None)
        if not clause_ref or verdict is None:
            continue
        why = _clean(values[3] if len(values) > 3 else None)
        if why:
            notes[clause_ref] = why
    return notes


def build_payload(workbook_path: Path) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover — script-only dependency
        raise SystemExit(
            "openpyxl is required to read the workbook. The checked-in payload at "
            "specs/standards/ is the source of truth for tests and seeding; this "
            "script only needs to run when the workbook is re-issued."
        ) from exc

    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    clause_rows = _extract_clause_rows(workbook[SHEET_CLAUSES])
    trap_notes = _extract_trap_notes(workbook[SHEET_TRAPS])

    for row in clause_rows:
        note = trap_notes.get(row["clause_ref"])
        if note and note != row.get("rationale"):
            row["trap_note"] = note

    return {
        "source_ref": "PEL-HSEQ-5064",
        "version_label": "1.0",
        "title": "Standards Alignment Matrix",
        "source_date": "2026-08-11",
        "excluded_frameworks": list(EXCLUDED_FRAMEWORKS),
        "notes": (
            "Annex SL clauses 4 to 10 across the five management system standards, "
            "plus the sheet 06 EXACT alignments for Investors in People and Annex A "
            "5.31 / 5.33. Sheet 03 (all 93 Annex A controls) and sheets 04/05 are not "
            "imported by this edition."
        ),
        "rows": clause_rows,
        "supplementary_rows": [dict(row) for row in SUPPLEMENTARY_ROWS],
    }


def main(argv: Optional[list[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument(
        "--out",
        type=Path,
        default=Path("specs/standards/pel-hseq-5064-alignment-v1.0.json"),
    )
    args = parser.parse_args(argv)

    if not args.workbook.exists():
        print(f"[extract_5064] workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    payload = build_payload(args.workbook)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(
        f"[extract_5064] wrote {args.out} — {len(payload['rows'])} clause rows, "
        f"{len(payload['supplementary_rows'])} supplementary rows"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
